#!/usr/bin/env python3
"""
–ï–î–ò–ù–´–ô –ü–ê–ô–ü–õ–ê–ô–ù: OCR ‚Üí Excel ‚Üí –°–≤–µ—Ä–∫–∞ —Å –ë–î ‚Üí –ò—Ç–æ–≥–æ–≤—ã–π –æ—Ç—á—ë—Ç.

–û–±—ä–µ–¥–∏–Ω—è–µ—Ç client_card_ocr.py –∏ verify_with_db.py
–≤ –æ–¥–∏–Ω –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏–π –ø—Ä–æ—Ü–µ—Å—Å.

–ü–æ—Ä—è–¥–æ–∫ —Ä–∞–±–æ—Ç—ã:
  –®–ê–ì 1 ‚Äî OCR + –ø–∞—Ä—Å–∏–Ω–≥ (Google Vision + Claude API)
  –®–ê–ì 2 ‚Äî –ì—Ä—É–ø–ø–∏—Ä–æ–≤–∫–∞ –∫–ª–∏–µ–Ω—Ç–æ–≤ (fuzzy matching –§–ò–û/–ò–ò–ù/—Ç–µ–ª–µ—Ñ–æ–Ω)
  –®–ê–ì 3 ‚Äî –î–µ–¥—É–ø–ª–∏–∫–∞—Ü–∏—è —Å—Ç—Ä–∞–Ω–∏—Ü (–ø–µ—Ä—Ü–µ–ø—Ç–∏–≤–Ω—ã–π —Ö—ç—à + OCR-—Ç–µ–∫—Å—Ç)
  –®–ê–ì 4 ‚Äî –ó–∞–ø–∏—Å—å –∫–ª–∏–µ–Ω—Ç—Å–∫–æ–π –±–∞–∑—ã –≤ Excel (6 –ª–∏—Å—Ç–æ–≤)
  –®–ê–ì 5 ‚Äî –°–≤–µ—Ä–∫–∞ —Å –ë–î ¬´–ü—Ä–∏–≤–∏–ª–µ–≥–∏—è¬ª (–º–∞—Ç—á–∏–Ω–≥ –§–ò–û + —Ç–µ–ª–µ—Ñ–æ–Ω)
  –®–ê–ì 6 ‚Äî –ò—Ç–æ–≥–æ–≤—ã–π –æ—Ç—á—ë—Ç (Excel + –∫–æ–Ω—Å–æ–ª—å)

–ó–∞–ø—É—Å–∫:
    python run_pipeline.py              # –ü–æ–ª–Ω—ã–π –ø–∞–π–ø–ª–∞–π–Ω
    python run_pipeline.py --skip-ocr   # –¢–æ–ª—å–∫–æ —Å–≤–µ—Ä–∫–∞ (OCR —É–∂–µ —Å–¥–µ–ª–∞–Ω)
    python run_pipeline.py --only-ocr   # –¢–æ–ª—å–∫–æ OCR (–±–µ–∑ —Å–≤–µ—Ä–∫–∏)

–†–µ–∑—É–ª—å—Ç–∞—Ç—ã:
    clients_database.xlsx      ‚Äî –∫–ª–∏–µ–Ω—Ç—Å–∫–∞—è –±–∞–∑–∞ (6 –ª–∏—Å—Ç–æ–≤)
    verification_report.xlsx   ‚Äî –æ—Ç—á—ë—Ç —Å–≤–µ—Ä–∫–∏ —Å –ë–î
    pipeline_report.xlsx       ‚Äî –∏—Ç–æ–≥–æ–≤—ã–π –∫–æ–º–±–∏–Ω–∏—Ä–æ–≤–∞–Ω–Ω—ã–π –æ—Ç—á—ë—Ç
    raw_results.json           ‚Äî —Å—ã—Ä—ã–µ –¥–∞–Ω–Ω—ã–µ OCR
    ocr_logs/ocr_YYYY-MM-DD.log ‚Äî –¥–µ—Ç–∞–ª—å–Ω—ã–π –ª–æ–≥
"""

import os
import sys

# –§–∏–∫—Å SSL –¥–ª—è gRPC –Ω–∞ macOS: —É–∫–∞–∑—ã–≤–∞–µ–º –ø—É—Ç—å –∫ –∫–æ—Ä–Ω–µ–≤—ã–º —Å–µ—Ä—Ç–∏—Ñ–∏–∫–∞—Ç–∞–º
if not os.environ.get("GRPC_DEFAULT_SSL_ROOTS_FILE_PATH"):
    try:
        import certifi
        os.environ["GRPC_DEFAULT_SSL_ROOTS_FILE_PATH"] = certifi.where()
    except ImportError:
        pass

import time
import json
import argparse
import logging
from datetime import datetime

import pandas as pd

# ============================================================
# –ü—Ä–æ–≤–µ—Ä–∫–∞ –æ–∫—Ä—É–∂–µ–Ω–∏—è –ø–µ—Ä–µ–¥ –∏–º–ø–æ—Ä—Ç–æ–º —Ç—è–∂—ë–ª—ã—Ö –º–æ–¥—É–ª–µ–π
# ============================================================

def check_config():
    """–ü—Ä–æ–≤–µ—Ä—è–µ—Ç –Ω–∞–ª–∏—á–∏–µ config.py –∏ –∫—Ä–∏—Ç–∏—á–µ—Å–∫–∏—Ö –ø–∞—Ä–∞–º–µ—Ç—Ä–æ–≤."""
    try:
        import config
    except ImportError:
        print("–û–®–ò–ë–ö–ê: –Ω–µ –Ω–∞–π–¥–µ–Ω config.py.")
        print("–ó–∞–ø—É—Å—Ç–∏—Ç–µ —Å–∫—Ä–∏–ø—Ç –∏–∑ –ø–∞–ø–∫–∏ ocr_project/")
        sys.exit(1)
    return config


def check_dependencies():
    """–ü—Ä–æ–≤–µ—Ä—è–µ—Ç –≤—Å–µ –∑–∞–≤–∏—Å–∏–º–æ—Å—Ç–∏."""
    missing = []
    for module in ['google.cloud.vision', 'anthropic', 'openpyxl', 'PIL', 'tqdm']:
        try:
            __import__(module)
        except ImportError:
            missing.append(module)
    if missing:
        print(f"–û–®–ò–ë–ö–ê: –Ω–µ —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω—ã –ø–∞–∫–µ—Ç—ã: {', '.join(missing)}")
        print("–£—Å—Ç–∞–Ω–æ–≤–∏—Ç–µ: pip install google-cloud-vision anthropic openpyxl Pillow tqdm")
        sys.exit(1)


# ============================================================
# –õ–û–ì–ò–†–û–í–ê–ù–ò–ï –î–õ–Ø –ü–ê–ô–ü–õ–ê–ô–ù–ê
# ============================================================

def setup_pipeline_logging(config):
    """–ù–∞—Å—Ç—Ä–æ–π–∫–∞ –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏—è –¥–ª—è –ø–∞–π–ø–ª–∞–π–Ω–∞."""
    log_folder = getattr(config, 'LOG_FOLDER', './ocr_logs')
    # Fallback: –µ—Å–ª–∏ –ø—É—Ç—å –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω, –ø–∏—à–µ–º –ª–æ–≥–∏ —Ä—è–¥–æ–º —Å–æ —Å–∫—Ä–∏–ø—Ç–æ–º
    try:
        os.makedirs(log_folder, exist_ok=True)
    except (PermissionError, OSError):
        log_folder = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "ocr_logs"
        )
        os.makedirs(log_folder, exist_ok=True)

    today = datetime.now().strftime('%Y-%m-%d')
    log_file = os.path.join(log_folder, f"pipeline_{today}.log")

    logger = logging.getLogger('pipeline')
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    # –§–∞–π–ª ‚Äî –ø–æ–ª–Ω—ã–π –ª–æ–≥
    fh = logging.FileHandler(log_file, encoding='utf-8')
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter(
        '%(asctime)s | %(levelname)-7s | %(message)s',
        datefmt='%H:%M:%S'
    ))
    logger.addHandler(fh)

    # –ö–æ–Ω—Å–æ–ª—å ‚Äî –æ—Å–Ω–æ–≤–Ω–æ–π –≤—ã–≤–æ–¥
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter('%(message)s'))
    logger.addHandler(ch)

    logger.info(f"–õ–æ–≥ –ø–∞–π–ø–ª–∞–π–Ω–∞: {log_file}")
    return logger


# ============================================================
# –®–ê–ì 1-4: OCR –ü–ê–ô–ü–õ–ê–ô–ù
# (–∏–º–ø–æ—Ä—Ç –∏–∑ client_card_ocr.py)
# ============================================================

def run_ocr_pipeline(log, config):
    """
    –ó–∞–ø—É—Å–∫–∞–µ—Ç –ø–æ–ª–Ω—ã–π OCR-–ø–∞–π–ø–ª–∞–π–Ω:
    Vision API ‚Üí Claude ‚Üí –≥—Ä—É–ø–ø–∏—Ä–æ–≤–∫–∞ ‚Üí –¥–µ–¥—É–ø–ª–∏–∫–∞—Ü–∏—è ‚Üí Excel.

    –í–æ–∑–≤—Ä–∞—â–∞–µ—Ç –ø—É—Ç—å –∫ —Å–æ–∑–¥–∞–Ω–Ω–æ–º—É Excel-—Ñ–∞–π–ª—É.
    """
    log.info("")
    log.info("‚ïî‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïó")
    log.info("‚ïë  –®–ê–ì 1-4: –û–¶–ò–§–†–û–í–ö–ê –ö–ê–†–¢–û–ß–ï–ö (OCR)                 ‚ïë")
    log.info("‚ïö‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïù")

    # –ò–º–ø–æ—Ä—Ç–∏—Ä—É–µ–º —Ñ—É–Ω–∫—Ü–∏–∏ –∏–∑ client_card_ocr
    try:
        from client_card_ocr import (
            init_vision_client, init_claude_client,
            process_all_images, group_by_client,
            deduplicate_pages, write_to_excel
        )
    except ImportError as e:
        log.error(f"–ù–µ —É–¥–∞–ª–æ—Å—å –∏–º–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞—Ç—å client_card_ocr: {e}")
        log.error("–£–±–µ–¥–∏—Ç–µ—Å—å, —á—Ç–æ client_card_ocr.py –Ω–∞—Ö–æ–¥–∏—Ç—Å—è –≤ —Ç–æ–π –∂–µ –ø–∞–ø–∫–µ.")
        return None

    # --- –ü—Ä–æ–≤–µ—Ä–∫–∏ ---
    if not os.path.exists(config.GOOGLE_VISION_CREDENTIALS):
        log.error(f"–§–∞–π–ª Google Vision –Ω–µ –Ω–∞–π–¥–µ–Ω: {config.GOOGLE_VISION_CREDENTIALS}")
        return None

    if not os.path.exists(config.INPUT_FOLDER):
        log.error(f"–ü–∞–ø–∫–∞ —Å —Ñ–æ—Ç–æ –Ω–µ –Ω–∞–π–¥–µ–Ω–∞: {config.INPUT_FOLDER}")
        return None

    # –°—á–∏—Ç–∞–µ–º —Ñ–æ—Ç–æ
    from pathlib import Path
    image_files = [
        f for f in os.listdir(config.INPUT_FOLDER)
        if Path(f).suffix.lower() in config.IMAGE_EXTENSIONS
    ]

    if not image_files:
        log.error(f"–§–æ—Ç–æ–≥—Ä–∞—Ñ–∏–∏ –Ω–µ –Ω–∞–π–¥–µ–Ω—ã –≤: {config.INPUT_FOLDER}")
        return None

    log.info(f"\n  –ü–∞–ø–∫–∞ —Ñ–æ—Ç–æ: {config.INPUT_FOLDER}")
    log.info(f"  –ù–∞–π–¥–µ–Ω–æ —Ñ–æ—Ç–æ: {len(image_files)}")
    log.info(f"  –ú–æ–¥–µ–ª—å Claude: {config.CLAUDE_MODEL}")

    base_url = getattr(config, 'ANTHROPIC_BASE_URL', None)
    if base_url:
        log.info(f"  Proxy: {base_url}")

    # --- –®–ê–ì 1: –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è API ---
    log.info("\n‚îÄ‚îÄ –®–ê–ì 1: –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è API ‚îÄ‚îÄ")
    try:
        vision_client = init_vision_client()
        log.info("  Google Vision ‚Äî OK ‚úì")
    except Exception as e:
        log.error(f"  Google Vision ‚Äî –û–®–ò–ë–ö–ê: {e}")
        return None

    try:
        claude_client = init_claude_client()
        log.info("  Claude API ‚Äî OK ‚úì")
    except Exception as e:
        log.error(f"  Claude API ‚Äî –û–®–ò–ë–ö–ê: {e}")
        return None

    # --- –®–ê–ì 2: OCR + –ø–∞—Ä—Å–∏–Ω–≥ ---
    log.info("\n‚îÄ‚îÄ –®–ê–ì 2: OCR + –∏–Ω—Ç–µ–ª–ª–µ–∫—Ç—É–∞–ª—å–Ω—ã–π –ø–∞—Ä—Å–∏–Ω–≥ ‚îÄ‚îÄ")
    t_ocr = time.time()
    results = process_all_images(vision_client, claude_client)
    ocr_time = time.time() - t_ocr

    # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Å—ã—Ä—ã–µ –¥–∞–Ω–Ω—ã–µ
    raw_path = os.path.join(
        os.path.dirname(config.OUTPUT_FILE) or '.', "raw_results.json"
    )
    with open(raw_path, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    log.info(f"  –°—ã—Ä—ã–µ –¥–∞–Ω–Ω—ã–µ: {raw_path}")

    # –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ OCR
    page_types = {}
    errors = 0
    for r in results:
        pt = r.get("page_type", "unknown")
        page_types[pt] = page_types.get(pt, 0) + 1
        if pt == "error":
            errors += 1

    log.info(f"\n  –†–µ–∑—É–ª—å—Ç–∞—Ç—ã OCR ({ocr_time:.0f}—Å):")
    for pt, cnt in sorted(page_types.items()):
        log.info(f"    {pt}: {cnt}")
    if errors:
        log.warning(f"    ‚ö† –û—à–∏–±–æ–∫: {errors}")

    # --- –®–ê–ì 3: –ì—Ä—É–ø–ø–∏—Ä–æ–≤–∫–∞ ---
    log.info("\n‚îÄ‚îÄ –®–ê–ì 3: –ì—Ä—É–ø–ø–∏—Ä–æ–≤–∫–∞ –ø–æ –∫–ª–∏–µ–Ω—Ç–∞–º ‚îÄ‚îÄ")
    grouped = group_by_client(results)
    n_clients = len([k for k in grouped if k != '_unmatched'])
    log.info(f"  –£–Ω–∏–∫–∞–ª—å–Ω—ã—Ö –∫–ª–∏–µ–Ω—Ç–æ–≤: {n_clients}")

    # --- –®–ê–ì 4: –î–µ–¥—É–ø–ª–∏–∫–∞—Ü–∏—è + Excel ---
    log.info("\n‚îÄ‚îÄ –®–ê–ì 4: –î–µ–¥—É–ø–ª–∏–∫–∞—Ü–∏—è + –∑–∞–ø–∏—Å—å Excel ‚îÄ‚îÄ")
    grouped = deduplicate_pages(grouped)
    write_to_excel(grouped, results)

    log.info(f"\n  ‚úì Excel —Å–æ—Ö—Ä–∞–Ω—ë–Ω: {config.OUTPUT_FILE}")

    return config.OUTPUT_FILE


# ============================================================
# –®–ê–ì 4.5: –ù–û–†–ú–ê–õ–ò–ó–ê–¶–ò–Ø OCR ‚Üí –§–û–†–ú–ê–¢ –ë–î
# (–∏–º–ø–æ—Ä—Ç –∏–∑ normalize_ocr.py)
# ============================================================

def run_normalization(log, config, ocr_excel_path):
    """
    –ù–æ—Ä–º–∞–ª–∏–∑—É–µ—Ç OCR Excel ‚Äî –ø–µ—Ä–µ–∏–º–µ–Ω–æ–≤—ã–≤–∞–µ—Ç –ø–æ–ª—è –ø–æ–¥ —Ñ–æ—Ä–º–∞—Ç –ë–î,
    —Å–æ–∑–¥–∞—ë—Ç —Å–≤–æ–¥–Ω—ã–π –ª–∏—Å—Ç ¬´–í—Å–µ_–≤–∏–∑–∏—Ç—ã¬ª.

    –í–æ–∑–≤—Ä–∞—â–∞–µ—Ç –ø—É—Ç—å –∫ –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω–æ–º—É —Ñ–∞–π–ª—É –∏–ª–∏ None.
    """
    log.info("")
    log.info("‚ïî‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïó")
    log.info("‚ïë  –®–ê–ì 4.5: –ù–û–†–ú–ê–õ–ò–ó–ê–¶–ò–Ø OCR ‚Üí –§–û–†–ú–ê–¢ –ë–î             ‚ïë")
    log.info("‚ïö‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïù")

    try:
        from normalize_ocr import normalize_ocr_file
    except ImportError as e:
        log.error(f"–ù–µ —É–¥–∞–ª–æ—Å—å –∏–º–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞—Ç—å normalize_ocr: {e}")
        log.warning("–ù–æ—Ä–º–∞–ª–∏–∑–∞—Ü–∏—è –ø—Ä–æ–ø—É—â–µ–Ω–∞.")
        return None

    if not ocr_excel_path or not os.path.exists(ocr_excel_path):
        log.warning(f"  OCR-—Ñ–∞–π–ª –Ω–µ –Ω–∞–π–¥–µ–Ω: {ocr_excel_path}")
        log.warning("  –ù–æ—Ä–º–∞–ª–∏–∑–∞—Ü–∏—è –ø—Ä–æ–ø—É—â–µ–Ω–∞.")
        return None

    script_dir = os.path.dirname(os.path.abspath(__file__))
    normalized_name = getattr(config, 'NORMALIZED_FILE', 'clients_normalized.xlsx')
    normalized_path = os.path.join(script_dir, normalized_name)

    t_norm = time.time()
    result = normalize_ocr_file(ocr_excel_path, normalized_path)
    norm_time = time.time() - t_norm

    if result:
        log.info(f"  –ù–æ—Ä–º–∞–ª–∏–∑–∞—Ü–∏—è –∑–∞ {norm_time:.1f}—Å")
        return normalized_path
    else:
        log.warning("  –ù–æ—Ä–º–∞–ª–∏–∑–∞—Ü–∏—è –Ω–µ —É–¥–∞–ª–∞—Å—å.")
        return None


# ============================================================
# –§–õ–ê–ì –§–ò–ù–ê–õ–¨–ù–û–ô –í–ï–†–ò–§–ò–ö–ê–¶–ò–ò
# ============================================================

# –ó–Ω–∞—á–µ–Ω–∏—è env var / config, –∫–æ—Ç–æ—Ä—ã–µ –æ—Ç–∫–ª—é—á–∞—é—Ç —Ñ–∏–Ω–∞–ª—å–Ω—É—é –≤–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏—é
_FALSY_VERIF = {"false", "0", "no", "off"}
_TRUTHY_FLAGS = {"true", "1", "yes", "on"}


def _is_smoke_mode() -> bool:
    """True –µ—Å–ª–∏ SMOKE_MODE –∑–∞–¥–∞–Ω –∫–∞–∫ truthy (true/1/yes/on)."""
    return os.environ.get("SMOKE_MODE", "").lower().strip() in _TRUTHY_FLAGS


def _gsheets_disabled(cfg) -> bool:
    """
    –í–æ–∑–≤—Ä–∞—â–∞–µ—Ç True, –µ—Å–ª–∏ –≤—ã–≥—Ä—É–∑–∫–∞ –≤ Google Sheets –æ—Ç–∫–ª—é—á–µ–Ω–∞.

    –ü—Ä–∏–æ—Ä–∏—Ç–µ—Ç:
    1. SMOKE_MODE=true ‚Üí –≤—Å–µ–≥–¥–∞ –æ—Ç–∫–ª—é—á–µ–Ω–æ (—Ç–∏—Ö–∏–π –ø—Ä–æ–ø—É—Å–∫)
    2. ENV GSHEETS_UPLOAD_ENABLED=falsy ‚Üí –æ—Ç–∫–ª—é—á–µ–Ω–æ
    3. config.GSHEETS_UPLOAD_ENABLED=False ‚Üí –æ—Ç–∫–ª—é—á–µ–Ω–æ
    """
    if _is_smoke_mode():
        return True
    env_val = os.environ.get("GSHEETS_UPLOAD_ENABLED", "").lower().strip()
    if env_val:
        return env_val in _FALSY_VERIF
    return not bool(getattr(cfg, "GSHEETS_UPLOAD_ENABLED", False))


def _final_verification_disabled(cfg) -> bool:
    """
    –í–æ–∑–≤—Ä–∞—â–∞–µ—Ç True, –µ—Å–ª–∏ —Ñ–∏–Ω–∞–ª—å–Ω–∞—è –≤–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏—è Claude –æ—Ç–∫–ª—é—á–µ–Ω–∞.

    –ü—Ä–∏–æ—Ä–∏—Ç–µ—Ç:
    1. –ü–µ—Ä–µ–º–µ–Ω–Ω–∞—è –æ–∫—Ä—É–∂–µ–Ω–∏—è ENABLE_FINAL_VERIFICATION (–µ—Å–ª–∏ –∑–∞–¥–∞–Ω–∞):
       –∑–Ω–∞—á–µ–Ω–∏—è false / 0 / no / off —Ç—Ä–∞–∫—Ç—É—é—Ç—Å—è –∫–∞–∫ ¬´–≤—ã–∫–ª—é—á–µ–Ω–æ¬ª.
    2. config.ENABLE_FINAL_VERIFICATION (–µ—Å–ª–∏ ENV –Ω–µ –∑–∞–¥–∞–Ω–∞).
    """
    env_val = os.environ.get("ENABLE_FINAL_VERIFICATION", "").lower().strip()
    if env_val:                             # ENV –∑–∞–¥–∞–Ω–∞ ‚Üí –æ–Ω–∞ –ø—Ä–∏–æ—Ä–∏—Ç–µ—Ç–Ω–∞
        return env_val in _FALSY_VERIF
    # ENV –Ω–µ –∑–∞–¥–∞–Ω–∞ ‚Üí –±–µ—Ä—ë–º –∏–∑ config
    return not bool(getattr(cfg, "ENABLE_FINAL_VERIFICATION", True))


# ============================================================
# –®–ê–ì 5-6: –°–í–ï–†–ö–ê –° –ë–î
# (–∏–º–ø–æ—Ä—Ç –∏–∑ verify_with_db.py)
# ============================================================

def run_verification(log, config, ocr_excel_path):
    """
    –ó–∞–ø—É—Å–∫–∞–µ—Ç —Å–≤–µ—Ä–∫—É –æ—Ü–∏—Ñ—Ä–æ–≤–∞–Ω–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö —Å –ë–î ¬´–ü—Ä–∏–≤–∏–ª–µ–≥–∏—è¬ª.

    –í–æ–∑–≤—Ä–∞—â–∞–µ—Ç (–ø—É—Ç—å_–∫_–æ—Ç—á—ë—Ç—É, DataFrame_—Å–≤–µ—Ä–∫–∏) –∏–ª–∏ (None, None).
    """
    log.info("")
    log.info("‚ïî‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïó")
    log.info("‚ïë  –®–ê–ì 5-6: –°–í–ï–†–ö–ê –° –ë–î ¬´–ü–†–ò–í–ò–õ–ï–ì–ò–Ø¬ª                 ‚ïë")
    log.info("‚ïö‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïù")

    # –ò–º–ø–æ—Ä—Ç–∏—Ä—É–µ–º —Ñ—É–Ω–∫—Ü–∏–∏ –∏–∑ verify_with_db
    try:
        from verify_with_db import (
            load_db, load_ocr, build_db_client_index,
            verify_clients, generate_report as generate_verification_report,
            save_not_found_clients
        )
    except ImportError as e:
        log.error(f"–ù–µ —É–¥–∞–ª–æ—Å—å –∏–º–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞—Ç—å verify_with_db: {e}")
        log.error("–£–±–µ–¥–∏—Ç–µ—Å—å, —á—Ç–æ verify_with_db.py –Ω–∞—Ö–æ–¥–∏—Ç—Å—è –≤ —Ç–æ–π –∂–µ –ø–∞–ø–∫–µ.")
        return None, None

    # –û–ø—Ä–µ–¥–µ–ª—è–µ–º –ø—É—Ç–∏
    script_dir = os.path.dirname(os.path.abspath(__file__))
    db_path = os.path.join(script_dir, "db_privilage.xlsx")
    report_path = os.path.join(script_dir, "verification_report.xlsx")

    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞–ª–∏—á–∏–µ –ë–î
    if not os.path.exists(db_path):
        log.warning(f"  –ë–î –Ω–µ –Ω–∞–π–¥–µ–Ω–∞: {db_path}")
        log.warning("  –°–∫–æ–ø–∏—Ä—É–π—Ç–µ db_privilage.xlsx –≤ –ø–∞–ø–∫—É ocr_project/")
        log.warning("  –°–≤–µ—Ä–∫–∞ –ø—Ä–æ–ø—É—â–µ–Ω–∞.")
        return None, None

    # --- –ó–∞–≥—Ä—É–∑–∫–∞ –ë–î ---
    log.info("\n‚îÄ‚îÄ –®–ê–ì 5: –ó–∞–≥—Ä—É–∑–∫–∞ –∏ –∏–Ω–¥–µ–∫—Å–∞—Ü–∏—è –ë–î ‚îÄ‚îÄ")
    t_db = time.time()
    db_df = load_db(db_path)
    db_index = build_db_client_index(db_df)
    db_time = time.time() - t_db
    log.info(f"  –ö–ª–∏–µ–Ω—Ç–æ–≤ –≤ –ë–î: {len(db_index)}")
    log.info(f"  –ó–∞–≥—Ä—É–∑–∫–∞ –∑–∞ {db_time:.1f}—Å")

    # --- –°–≤–µ—Ä–∫–∞ ---
    log.info("\n‚îÄ‚îÄ –®–ê–ì 6: –ú–∞—Ç—á–∏–Ω–≥ OCR ‚Üî –ë–î ‚îÄ‚îÄ")

    threshold = getattr(config, 'DB_MATCH_THRESHOLD', 0.70)
    log.info(f"  –ü–æ—Ä–æ–≥ —Å–æ–≤–ø–∞–¥–µ–Ω–∏—è: {threshold*100:.0f}%")

    verification_df = pd.DataFrame()
    ocr_sheets = None

    if ocr_excel_path and os.path.exists(ocr_excel_path):
        t_match = time.time()
        ocr_sheets = load_ocr(ocr_excel_path)
        if ocr_sheets:
            verification_df = verify_clients(ocr_sheets, db_index, threshold)
        match_time = time.time() - t_match
        log.info(f"  –°–≤–µ—Ä–∫–∞ –∑–∞ {match_time:.1f}—Å")
    else:
        log.warning(f"  OCR-—Ñ–∞–π–ª –Ω–µ –Ω–∞–π–¥–µ–Ω: {ocr_excel_path}")
        log.warning("  –ì–µ–Ω–µ—Ä–∞—Ü–∏—è –æ—Ç—á—ë—Ç–∞ —Ç–æ–ª—å–∫–æ –∏–∑ –ë–î.")

    # ========== –®–ê–ì 6.5: –§–ò–ù–ê–õ–¨–ù–ê–Ø –í–ï–†–ò–§–ò–ö–ê–¶–ò–Ø CLAUDE ==========
    # Guard: ENV ENABLE_FINAL_VERIFICATION (–ø—Ä–∏–æ—Ä–∏—Ç–µ—Ç) –∏–ª–∏ config.ENABLE_FINAL_VERIFICATION.
    # –ó–Ω–∞—á–µ–Ω–∏—è false/0/no/off –æ—Ç–∫–ª—é—á–∞—é—Ç Claude. quality_baseline.py —Å—Ç–∞–≤–∏—Ç ENV=false.
    if _final_verification_disabled(config) and len(verification_df) > 0 and ocr_sheets:
        log.info("\n‚îÄ‚îÄ –®–ê–ì 6.5: –§–∏–Ω–∞–ª—å–Ω–∞—è –≤–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏—è –ø—Ä–æ–ø—É—â–µ–Ω–∞ (–æ—Ç–∫–ª—é—á–µ–Ω–æ) ‚îÄ‚îÄ")
    elif len(verification_df) > 0 and ocr_sheets:
        log.info("\n‚îÄ‚îÄ –®–ê–ì 6.5: –§–∏–Ω–∞–ª—å–Ω–∞—è –≤–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏—è Claude ‚îÄ‚îÄ")

        try:
            from final_verification import run_final_claude_verification
            from client_card_ocr import init_claude_client

            # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Ä–µ–∂–∏–º fallback-only
            fallback_only = getattr(config, 'FINAL_VERIFICATION_FALLBACK_ONLY', True)

            if fallback_only:
                # –ò–º–ø–æ—Ä—Ç–∏—Ä—É–µ–º –∫–æ–Ω—Å—Ç–∞–Ω—Ç—ã —Å—Ç–∞—Ç—É—Å–æ–≤
                try:
                    from config import STATUS_DB_MAYBE, STATUS_DB_NOT_FOUND
                except ImportError:
                    STATUS_DB_MAYBE = "–í–æ–∑–º–æ–∂–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ –≤ –ë–î"
                    STATUS_DB_NOT_FOUND = "–ù–µ—Ç –≤ –ë–î (–Ω–æ–≤—ã–π –¥–ª—è –∫–∞—Ä—Ç–æ—Ç–µ–∫–∏)"

                # –ò—Å–ø–æ–ª—å–∑—É–µ–º –°—Ç–∞—Ç—É—Å_–ë–î –µ—Å–ª–∏ –¥–æ—Å—Ç—É–ø–µ–Ω, –∏–Ω–∞—á–µ –°—Ç–∞—Ç—É—Å (backward compatibility)
                status_column = "–°—Ç–∞—Ç—É—Å_–ë–î" if "–°—Ç–∞—Ç—É—Å_–ë–î" in verification_df.columns else "–°—Ç–∞—Ç—É—Å"

                # –§–∏–ª—å—Ç—Ä—É–µ–º —Å—Ç—Ä–æ–∫–∏ –¥–ª—è fallback-–≤–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏–∏
                if status_column == "–°—Ç–∞—Ç—É—Å_–ë–î":
                    # –ù–æ–≤–∞—è —Å–∏—Å—Ç–µ–º–∞: –±–µ—Ä—ë–º "–í–æ–∑–º–æ–∂–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ" –∏ "–ù–µ—Ç –≤ –ë–î"
                    fallback_mask = verification_df[status_column].isin([
                        STATUS_DB_MAYBE,
                        STATUS_DB_NOT_FOUND
                    ])
                    status_names = f"{STATUS_DB_MAYBE} / {STATUS_DB_NOT_FOUND}"
                else:
                    # –°—Ç–∞—Ä–∞—è —Å–∏—Å—Ç–µ–º–∞ (backward compatibility)
                    fallback_mask = verification_df[status_column].isin(["–ù–µ –Ω–∞–π–¥–µ–Ω", "–í–æ–∑–º–æ–∂–Ω–æ"])
                    status_names = "–ù–µ –Ω–∞–π–¥–µ–Ω / –í–æ–∑–º–æ–∂–Ω–æ"

                fallback_df = verification_df[fallback_mask].copy()

                if len(fallback_df) == 0:
                    log.info(f"  Fallback-—Ä–µ–∂–∏–º: –Ω–µ—Ç –∫–ª–∏–µ–Ω—Ç–æ–≤ –¥–ª—è –≤–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏–∏ (–∫–æ–ª–æ–Ω–∫–∞: {status_column})")
                    log.info("  –§–∏–Ω–∞–ª—å–Ω–∞—è –≤–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏—è –ø—Ä–æ–ø—É—â–µ–Ω–∞.")
                else:
                    log.info(f"  Fallback-—Ä–µ–∂–∏–º: –≤–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏—è {len(fallback_df)} –∫–ª–∏–µ–Ω—Ç–æ–≤")
                    log.info(f"  –°—Ç–∞—Ç—É—Å—ã: {status_names}")

                    # –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è Claude –∫–ª–∏–µ–Ω—Ç–∞
                    claude_client = init_claude_client()

                    t_verify = time.time()
                    enhanced_fallback_df, final_report_path = run_final_claude_verification(
                        log=log,
                        config=config,
                        claude_client=claude_client,
                        verification_df=fallback_df,
                        ocr_sheets=ocr_sheets,
                        db_index=db_index
                    )
                    verify_time = time.time() - t_verify

                    log.info(f"  Claude –≤–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏—è –∑–∞ {verify_time:.1f}—Å")
                    if final_report_path:
                        log.info(f"  ‚úì –û—Ç—á—ë—Ç –≤–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏–∏: {final_report_path}")

                    # –ú–µ—Ä–∂–∏–º —Ä–µ–∑—É–ª—å—Ç–∞—Ç –æ–±—Ä–∞—Ç–Ω–æ –≤ –ø–æ–ª–Ω—ã–π verification_df
                    # –û–±–Ω–æ–≤–ª—è–µ–º —Ç–æ–ª—å–∫–æ —Ç–µ —Å—Ç—Ä–æ–∫–∏, –∫–æ—Ç–æ—Ä—ã–µ –±—ã–ª–∏ –æ–±—Ä–∞–±–æ—Ç–∞–Ω—ã
                    for idx in enhanced_fallback_df.index:
                        if idx in verification_df.index:
                            for col in enhanced_fallback_df.columns:
                                if col.startswith('Claude_') or col in ['–í–æ–∑–º–æ–∂–Ω—ã–µ_—Å–æ–≤–ø–∞–¥–µ–Ω–∏—è_–ë–î', '–†–∞—Å—Ö–æ–∂–¥–µ–Ω–∏—è', '–†–µ–∫–æ–º–µ–Ω–¥–∞—Ü–∏–∏', '–ò—Å–ø—Ä–∞–≤–ª–µ–Ω–∏—è_OCR']:
                                    verification_df.at[idx, col] = enhanced_fallback_df.at[idx, col]
                        else:
                            # –ï—Å–ª–∏ –∏–Ω–¥–µ–∫—Å –Ω–µ —Å–æ–≤–ø–∞–¥–∞–µ—Ç, –¥–æ–±–∞–≤–ª—è–µ–º –Ω–æ–≤—É—é —Å—Ç—Ä–æ–∫—É (–Ω–∞ –≤—Å—è–∫–∏–π —Å–ª—É—á–∞–π)
                            verification_df = pd.concat([verification_df, enhanced_fallback_df.loc[[idx]]], ignore_index=False)

                    log.info(f"  –û–±–Ω–æ–≤–ª–µ–Ω–æ {len(fallback_df)} –∑–∞–ø–∏—Å–µ–π –≤ verification_df")

            else:
                # –†–µ–∂–∏–º "–≤—Å–µ –∫–ª–∏–µ–Ω—Ç—ã"
                log.info(f"  –ü–æ–ª–Ω–∞—è –≤–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏—è –≤—Å–µ—Ö {len(verification_df)} –∫–ª–∏–µ–Ω—Ç–æ–≤")

                # –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è Claude –∫–ª–∏–µ–Ω—Ç–∞
                claude_client = init_claude_client()

                t_verify = time.time()
                enhanced_verification_df, final_report_path = run_final_claude_verification(
                    log=log,
                    config=config,
                    claude_client=claude_client,
                    verification_df=verification_df,
                    ocr_sheets=ocr_sheets,
                    db_index=db_index
                )
                verify_time = time.time() - t_verify

                log.info(f"  Claude –≤–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏—è –∑–∞ {verify_time:.1f}—Å")
                if final_report_path:
                    log.info(f"  ‚úì –û—Ç—á—ë—Ç –≤–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏–∏: {final_report_path}")

                # –ò—Å–ø–æ–ª—å–∑—É–µ–º –æ–±–æ–≥–∞—â—ë–Ω–Ω—ã–π DataFrame –¥–ª—è –¥–∞–ª—å–Ω–µ–π—à–µ–π —Ä–∞–±–æ—Ç—ã
                verification_df = enhanced_verification_df

        except ImportError as e:
            log.warning(f"  –ú–æ–¥—É–ª—å final_verification –Ω–µ –Ω–∞–π–¥–µ–Ω: {e}")
            log.warning("  –§–∏–Ω–∞–ª—å–Ω–∞—è –≤–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏—è –ø—Ä–æ–ø—É—â–µ–Ω–∞.")
        except Exception as e:
            log.error(f"  –û—à–∏–±–∫–∞ —Ñ–∏–Ω–∞–ª—å–Ω–æ–π –≤–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏–∏: {e}")
            log.warning("  –ü—Ä–æ–¥–æ–ª–∂–∞–µ–º –±–µ–∑ —Ñ–∏–Ω–∞–ª—å–Ω–æ–π –≤–µ—Ä–∏—Ñ–∏–∫–∞—Ü–∏–∏.")
    # ============================================================

    # --- –ì–µ–Ω–µ—Ä–∞—Ü–∏—è –æ—Ç—á—ë—Ç–∞ ---
    generate_verification_report(verification_df, db_df, report_path)
    log.info(f"  ‚úì –û—Ç—á—ë—Ç —Å–≤–µ—Ä–∫–∏: {report_path}")

    # --- –°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ –Ω–µ–Ω–∞–π–¥–µ–Ω–Ω—ã—Ö –∫–ª–∏–µ–Ω—Ç–æ–≤ ---
    if len(verification_df) > 0 and ocr_sheets:
        try:
            not_found_file = getattr(config, 'NOT_FOUND_CLIENTS_FILE', 'clients_not_found.xlsx')
        except:
            not_found_file = 'clients_not_found.xlsx'

        not_found_path = os.path.join(script_dir, not_found_file)
        save_not_found_clients(verification_df, ocr_sheets, not_found_path)

    return report_path, verification_df


# ============================================================
# –ò–¢–û–ì–û–í–´–ô –ö–û–ú–ë–ò–ù–ò–†–û–í–ê–ù–ù–´–ô –û–¢–ß–Å–¢
# ============================================================

def generate_pipeline_report(log, config, verification_df, ocr_excel_path):
    """
    –ì–µ–Ω–µ—Ä–∏—Ä—É–µ—Ç –∏—Ç–æ–≥–æ–≤—ã–π –∫–æ–º–±–∏–Ω–∏—Ä–æ–≤–∞–Ω–Ω—ã–π –æ—Ç—á—ë—Ç pipeline_report.xlsx
    —Å–æ –≤—Å–µ–º–∏ –∫–ª—é—á–µ–≤—ã–º–∏ –¥–∞–Ω–Ω—ã–º–∏ –≤ –æ–¥–Ω–æ–º —Ñ–∞–π–ª–µ.
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    report_path = os.path.join(script_dir, "pipeline_report.xlsx")

    log.info("\n‚îÄ‚îÄ –ì–µ–Ω–µ—Ä–∞—Ü–∏—è –∏—Ç–æ–≥–æ–≤–æ–≥–æ –æ—Ç—á—ë—Ç–∞ ‚îÄ‚îÄ")

    try:
        with pd.ExcelWriter(report_path, engine="openpyxl") as writer:

            # –õ–∏—Å—Ç 1: –°–≤–æ–¥–∫–∞ –ø–∞–π–ø–ª–∞–π–Ω–∞
            summary_data = {
                "–ü–∞—Ä–∞–º–µ—Ç—Ä": [
                    "–î–∞—Ç–∞ –∑–∞–ø—É—Å–∫–∞",
                    "–ú–æ–¥–µ–ª—å Claude",
                    "–ü–∞–ø–∫–∞ —Ñ–æ—Ç–æ",
                    "–ü–æ—Ä–æ–≥ –≥—Ä—É–ø–ø–∏—Ä–æ–≤–∫–∏ –§–ò–û",
                    "–ü–æ—Ä–æ–≥ –¥–µ–¥—É–ø–ª–∏–∫–∞—Ü–∏–∏ OCR",
                    "–ü–æ—Ä–æ–≥ —Å–≤–µ—Ä–∫–∏ —Å –ë–î",
                ],
                "–ó–Ω–∞—á–µ–Ω–∏–µ": [
                    datetime.now().strftime("%d.%m.%Y %H:%M"),
                    config.CLAUDE_MODEL,
                    config.INPUT_FOLDER,
                    f"{getattr(config, 'FUZZY_NAME_THRESHOLD', 0.75)*100:.0f}%",
                    f"{getattr(config, 'OCR_DUPLICATE_THRESHOLD', 0.90)*100:.0f}%",
                    f"{getattr(config, 'DB_MATCH_THRESHOLD', 0.70)*100:.0f}%",
                ]
            }
            pd.DataFrame(summary_data).to_excel(
                writer, sheet_name="–°–≤–æ–¥–∫–∞", index=False
            )

            # –õ–∏—Å—Ç 2: –†–µ–∑—É–ª—å—Ç–∞—Ç—ã —Å–≤–µ—Ä–∫–∏ (–µ—Å–ª–∏ –µ—Å—Ç—å)
            if verification_df is not None and len(verification_df) > 0:
                verification_df.to_excel(
                    writer, sheet_name="–°–≤–µ—Ä–∫–∞_OCR_vs_–ë–î", index=False
                )

                # –õ–∏—Å—Ç 3: –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ —Å–≤–µ—Ä–∫–∏
                status_col = "–°—Ç–∞—Ç—É—Å_–ë–î" if "–°—Ç–∞—Ç—É—Å_–ë–î" in verification_df.columns else "–°—Ç–∞—Ç—É—Å"
                stats = verification_df[status_col].value_counts().reset_index()
                stats.columns = ["–°—Ç–∞—Ç—É—Å", "–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ"]
                total = len(verification_df)
                stats["–î–æ–ª—è_%"] = (stats["–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ"] / total * 100).round(1)
                stats.to_excel(
                    writer, sheet_name="–°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞_—Å–≤–µ—Ä–∫–∏", index=False
                )

            # –õ–∏—Å—Ç 4: –ö–ª–∏–µ–Ω—Ç—ã –∏–∑ OCR (–µ—Å–ª–∏ Excel —Å—É—â–µ—Å—Ç–≤—É–µ—Ç)
            if ocr_excel_path and os.path.exists(ocr_excel_path):
                try:
                    ocr_clients = pd.read_excel(ocr_excel_path, sheet_name="–ö–ª–∏–µ–Ω—Ç—ã")
                    ocr_clients.to_excel(
                        writer, sheet_name="–ö–ª–∏–µ–Ω—Ç—ã_OCR", index=False
                    )
                except Exception:
                    pass

            # –õ–∏—Å—Ç 5: –ë–î –ü—Ä–∏–≤–∏–ª–µ–≥–∏—è ‚Äî —Ç–æ–ø –∫–ª–∏–µ–Ω—Ç—ã
            db_path = os.path.join(script_dir, "db_privilage.xlsx")
            if os.path.exists(db_path):
                try:
                    db_df = pd.read_excel(db_path)
                    db_df.columns = ["id", "name", "phone", "date",
                                     "doctor", "service", "qty"]

                    # –¢–æ–ø –∫–ª–∏–µ–Ω—Ç—ã
                    top = (
                        db_df.groupby("name")
                        .agg(–≤–∏–∑–∏—Ç–æ–≤=("name", "size"),
                             —Ç–µ–ª–µ—Ñ–æ–Ω=("phone", "first"))
                        .reset_index()
                        .sort_values("–≤–∏–∑–∏—Ç–æ–≤", ascending=False)
                        .head(100)
                    )
                    top.columns = ["–§–ò–û", "–í–∏–∑–∏—Ç–æ–≤", "–¢–µ–ª–µ—Ñ–æ–Ω"]
                    top.to_excel(
                        writer, sheet_name="–¢–æ–ø_–∫–ª–∏–µ–Ω—Ç—ã_–ë–î", index=False
                    )

                    # –í—Ä–∞—á–∏
                    doctors = db_df["doctor"].value_counts().reset_index()
                    doctors.columns = ["–í—Ä–∞—á", "–ó–∞–ø–∏—Å–µ–π"]
                    doctors.to_excel(
                        writer, sheet_name="–í—Ä–∞—á–∏_–ë–î", index=False
                    )

                    # –£—Å–ª—É–≥–∏
                    services = db_df["service"].value_counts().reset_index().head(50)
                    services.columns = ["–£—Å–ª—É–≥–∞", "–ö–æ–ª-–≤–æ"]
                    services.to_excel(
                        writer, sheet_name="–¢–æ–ø_—É—Å–ª—É–≥–∏_–ë–î", index=False
                    )
                except Exception as e:
                    log.warning(f"  –ù–µ —É–¥–∞–ª–æ—Å—å –ø—Ä–æ—á–∏—Ç–∞—Ç—å –ë–î: {e}")

        log.info(f"  ‚úì –ò—Ç–æ–≥–æ–≤—ã–π –æ—Ç—á—ë—Ç: {report_path}")
        return report_path

    except Exception as e:
        log.error(f"  –û—à–∏–±–∫–∞ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ –æ—Ç—á—ë—Ç–∞: {e}")
        return None


# ============================================================
# –ü–ï–ß–ê–¢–¨ –§–ò–ù–ê–õ–¨–ù–û–ô –°–í–û–î–ö–ò
# ============================================================

def print_summary(log, verification_df, total_time, ocr_excel_path, config):
    """–ö—Ä–∞—Å–∏–≤–∞—è —Å–≤–æ–¥–∫–∞ –≤ –∫–æ–Ω—Å–æ–ª—å."""
    log.info("")
    log.info("‚ïî‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïó")
    log.info("‚ïë              –ò–¢–û–ì–û–í–ê–Ø –°–í–û–î–ö–ê –ü–ê–ô–ü–õ–ê–ô–ù–ê              ‚ïë")
    log.info("‚ïö‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïù")

    # OCR —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞
    if ocr_excel_path and os.path.exists(ocr_excel_path):
        try:
            ocr_df = pd.read_excel(ocr_excel_path, sheet_name="–ö–ª–∏–µ–Ω—Ç—ã")
            log.info(f"\n  üìã –û–¶–ò–§–†–û–í–ö–ê:")
            log.info(f"     –ö–ª–∏–µ–Ω—Ç–æ–≤ —Ä–∞—Å–ø–æ–∑–Ω–∞–Ω–æ: {len(ocr_df)}")
        except Exception:
            pass

    # –°–≤–µ—Ä–∫–∞
    if verification_df is not None and len(verification_df) > 0:
        total = len(verification_df)

        # –ò—Å–ø–æ–ª—å–∑—É–µ–º –Ω–æ–≤—ã–µ —Å—Ç–∞—Ç—É—Å—ã (–°—Ç–∞—Ç—É—Å_–ë–î) –µ—Å–ª–∏ –¥–æ—Å—Ç—É–ø–Ω—ã, –∏–Ω–∞—á–µ —Å—Ç–∞—Ä—ã–µ
        status_column = "–°—Ç–∞—Ç—É—Å_–ë–î" if "–°—Ç–∞—Ç—É—Å_–ë–î" in verification_df.columns else "–°—Ç–∞—Ç—É—Å"

        # –ò–º–ø–æ—Ä—Ç–∏—Ä—É–µ–º –Ω–æ–≤—ã–µ —Å—Ç–∞—Ç—É—Å—ã
        try:
            from config import STATUS_DB_FOUND, STATUS_DB_MAYBE, STATUS_DB_NOT_FOUND
        except ImportError:
            STATUS_DB_FOUND = "–ù–∞–π–¥–µ–Ω –≤ –ë–î"
            STATUS_DB_MAYBE = "–í–æ–∑–º–æ–∂–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ –≤ –ë–î"
            STATUS_DB_NOT_FOUND = "–ù–µ—Ç –≤ –ë–î (–Ω–æ–≤—ã–π –¥–ª—è –∫–∞—Ä—Ç–æ—Ç–µ–∫–∏)"

        # –ü–æ–¥—Å—á–∏—Ç—ã–≤–∞–µ–º —Å —É—á–µ—Ç–æ–º –Ω–æ–≤—ã—Ö –∏ —Å—Ç–∞—Ä—ã—Ö —Å—Ç–∞—Ç—É—Å–æ–≤
        if status_column == "–°—Ç–∞—Ç—É—Å_–ë–î":
            found = len(verification_df[verification_df[status_column] == STATUS_DB_FOUND])
            maybe = len(verification_df[verification_df[status_column] == STATUS_DB_MAYBE])
            not_found = len(verification_df[verification_df[status_column] == STATUS_DB_NOT_FOUND])
        else:
            # Backward compatibility
            found = len(verification_df[verification_df[status_column] == "–ù–∞–π–¥–µ–Ω"])
            maybe = len(verification_df[verification_df[status_column] == "–í–æ–∑–º–æ–∂–Ω–æ"])
            not_found = len(verification_df[verification_df[status_column] == "–ù–µ –Ω–∞–π–¥–µ–Ω"])

        log.info(f"\n  üîç –°–í–ï–†–ö–ê –° –ë–î:")
        log.info(f"     –í—Å–µ–≥–æ –æ—Ü–∏—Ñ—Ä–æ–≤–∞–Ω–æ:         {total}")
        log.info(f"     ‚úì –ù–∞–π–¥–µ–Ω–æ –≤ –ë–î:           {found} ({found/total*100:.0f}%)")
        log.info(f"     ~ –í–æ–∑–º–æ–∂–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ:   {maybe} ({maybe/total*100:.0f}%)")
        log.info(f"     ‚äï –ù–æ–≤—ã–µ –¥–ª—è –∫–∞—Ä—Ç–æ—Ç–µ–∫–∏:    {not_found} ({not_found/total*100:.0f}%)")

    # –§–∞–π–ª—ã
    script_dir = os.path.dirname(os.path.abspath(__file__))
    log.info(f"\n  üìÅ –§–ê–ô–õ–´:")
    if ocr_excel_path:
        log.info(f"     –ö–ª–∏–µ–Ω—Ç—Å–∫–∞—è –±–∞–∑–∞:     {ocr_excel_path}")

    norm_name = getattr(config, 'NORMALIZED_FILE', 'clients_normalized.xlsx')
    norm_path = os.path.join(script_dir, norm_name)
    if os.path.exists(norm_path):
        log.info(f"     –ù–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω—ã–π:     {norm_path}")

    report_path = os.path.join(script_dir, "verification_report.xlsx")
    if os.path.exists(report_path):
        log.info(f"     –û—Ç—á—ë—Ç —Å–≤–µ—Ä–∫–∏:        {report_path}")

    pipeline_path = os.path.join(script_dir, "pipeline_report.xlsx")
    if os.path.exists(pipeline_path):
        log.info(f"     –ò—Ç–æ–≥–æ–≤—ã–π –æ—Ç—á—ë—Ç:      {pipeline_path}")

    not_found_file = getattr(config, 'NOT_FOUND_CLIENTS_FILE', 'clients_not_found.xlsx')
    not_found_path = os.path.join(script_dir, not_found_file)
    if os.path.exists(not_found_path):
        log.info(f"     –ù–µ –Ω–∞–π–¥–µ–Ω—ã –≤ –ë–î:     {not_found_path}")

    # –í—Ä–µ–º—è
    log.info(f"\n  ‚è±  –û–±—â–µ–µ –≤—Ä–µ–º—è: {total_time:.0f}—Å ({total_time/60:.1f} –º–∏–Ω)")
    log.info("")
    log.info("‚ïê" * 56)


# ============================================================
# MAIN
# ============================================================

def parse_args():
    parser = argparse.ArgumentParser(
        description="–ï–¥–∏–Ω—ã–π –ø–∞–π–ø–ª–∞–π–Ω: OCR ‚Üí Excel ‚Üí –°–≤–µ—Ä–∫–∞ —Å –ë–î ‚Üí –û—Ç—á—ë—Ç",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
–ü—Ä–∏–º–µ—Ä—ã:
  python run_pipeline.py              # –ü–æ–ª–Ω—ã–π –ø–∞–π–ø–ª–∞–π–Ω (–Ω–æ–≤—ã–µ –∫–∞—Ä—Ç–æ—á–∫–∏)
  python run_pipeline.py --skip-ocr   # –¢–æ–ª—å–∫–æ —Å–≤–µ—Ä–∫–∞ (OCR —É–∂–µ –≥–æ—Ç–æ–≤)
  python run_pipeline.py --only-ocr   # –¢–æ–ª—å–∫–æ OCR (–±–µ–∑ —Å–≤–µ—Ä–∫–∏)
  python run_pipeline.py --force      # –û–±—Ä–∞–±–æ—Ç–∞—Ç—å –í–°–ï –∑–∞–Ω–æ–≤–æ (—Å–±—Ä–æ—Å —Ä–µ–µ—Å—Ç—Ä–∞)
        """
    )
    parser.add_argument(
        '--skip-ocr', action='store_true',
        help='–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å OCR (–∏—Å–ø–æ–ª—å–∑–æ–≤–∞—Ç—å —Å—É—â–µ—Å—Ç–≤—É—é—â–∏–π clients_database.xlsx)'
    )
    parser.add_argument(
        '--only-ocr', action='store_true',
        help='–¢–æ–ª—å–∫–æ OCR, –±–µ–∑ —Å–≤–µ—Ä–∫–∏ —Å –ë–î'
    )
    parser.add_argument(
        '--force', action='store_true',
        help='–°–±—Ä–æ—Å–∏—Ç—å —Ä–µ–µ—Å—Ç—Ä –∏ –æ–±—Ä–∞–±–æ—Ç–∞—Ç—å –í–°–ï –∫–∞—Ä—Ç–æ—á–∫–∏ –∑–∞–Ω–æ–≤–æ'
    )
    return parser.parse_args()


def add_verification_sheet(clients_path: str, verification_df: pd.DataFrame, log: logging.Logger):
    """
    –î–æ–±–∞–≤–ª—è–µ—Ç –ª–∏—Å—Ç ¬´–°–≤–µ—Ä–∫–∞_–ë–î¬ª –≤ clients_database.xlsx
    —Å —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞–º–∏ –±—ã—Å—Ç—Ä–æ–π —Å–≤–µ—Ä–∫–∏ OCR-–∫–ª–∏–µ–Ω—Ç–æ–≤ —Å –ë–î ¬´–ü—Ä–∏–≤–∏–ª–µ–≥–∏—è¬ª.
    """
    if not os.path.exists(clients_path):
        log.warning(f"  add_verification_sheet: —Ñ–∞–π–ª –Ω–µ –Ω–∞–π–¥–µ–Ω: {clients_path}")
        return
    if verification_df is None or len(verification_df) == 0:
        log.warning("  add_verification_sheet: verification_df –ø—É—Å—Ç–æ–π")
        return

    from openpyxl import load_workbook
    from zipfile import BadZipFile

    # –ö–æ–ª–æ–Ω–∫–∏ –¥–ª—è –ª–∏—Å—Ç–∞ —Å–≤–µ—Ä–∫–∏ (—Ç–æ–ª—å–∫–æ –∫–ª—é—á–µ–≤—ã–µ, –±–µ–∑ OCR-—Ç–µ–∫—Å—Ç–æ–≤)
    keep_cols = [
        "OCR_–§–ò–û", "OCR_–¢–µ–ª–µ—Ñ–æ–Ω",
        "–°—Ç–∞—Ç—É—Å_–ë–î", "–ë–î_ID",
        "–ë–î_–§–ò–û", "–ë–î_–¢–µ–ª–µ—Ñ–æ–Ω",
        "–°–æ–≤–ø–∞–¥–µ–Ω–∏–µ_%", "–í–∏–∑–∏—Ç–æ–≤_–≤_–ë–î", "–í—Ä–∞—á–∏_–≤_–ë–î",
    ]
    # –û—Å—Ç–∞–≤–ª—è–µ–º —Ç–æ–ª—å–∫–æ —Å—É—â–µ—Å—Ç–≤—É—é—â–∏–µ –∫–æ–ª–æ–Ω–∫–∏
    cols = [c for c in keep_cols if c in verification_df.columns]
    vdf = verification_df[cols].copy()

    try:
        wb = load_workbook(clients_path)
    except BadZipFile:
        log.warning(f"  ‚ö† add_verification_sheet: —Ñ–∞–π–ª –ø–æ–≤—Ä–µ–∂–¥—ë–Ω (BadZipFile): {clients_path}")
        return
    except Exception as e:
        log.warning(f"  ‚ö† add_verification_sheet: –Ω–µ —É–¥–∞–ª–æ—Å—å –æ—Ç–∫—Ä—ã—Ç—å —Ñ–∞–π–ª: {e}")
        return

    # –£–¥–∞–ª—è–µ–º —Å—Ç–∞—Ä—ã–π –ª–∏—Å—Ç, –µ—Å–ª–∏ –µ—Å—Ç—å
    sheet_name = "–°–≤–µ—Ä–∫–∞_–ë–î"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    # –ó–∞–≥–æ–ª–æ–≤–∫–∏
    for col_idx, col_name in enumerate(vdf.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # –î–∞–Ω–Ω—ã–µ
    for row_idx, (_, row) in enumerate(vdf.iterrows(), 2):
        for col_idx, col_name in enumerate(vdf.columns, 1):
            val = row[col_name]
            if pd.isna(val):
                val = ""
            ws.cell(row=row_idx, column=col_idx, value=val)

    # –ê–≤—Ç–æ—Ñ–∏–ª—å—Ç—Ä
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions

    wb.save(clients_path)
    wb.close()
    log.info(f"  ‚úì –õ–∏—Å—Ç ¬´{sheet_name}¬ª –¥–æ–±–∞–≤–ª–µ–Ω –≤ {clients_path} ({len(vdf)} –∑–∞–ø–∏—Å–µ–π)")


def enrich_clients_with_db_match(clients_path: str, verification_df: pd.DataFrame, log: logging.Logger):
    """
    –î–æ–ø–æ–ª–Ω—è–µ—Ç clients_database.xlsx –∫–æ–ª–æ–Ω–∫–∞–º–∏ —Å –Ω–∞–π–¥–µ–Ω–Ω—ã–º –§–ò–û –∏–∑ –ë–î –∏ —Å—Ç–∞—Ç—É—Å–æ–º —Å–æ–≤–ø–∞–¥–µ–Ω–∏—è.
    –ö–æ–ª–æ–Ω–∫–∏ –¥–æ–±–∞–≤–ª—è—é—Ç—Å—è —Ç–æ–ª—å–∫–æ –≤ –ª–∏—Å—Ç '–ö–ª–∏–µ–Ω—Ç—ã': –ë–î_–§–ò–û_—Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ, –°—Ç–∞—Ç—É—Å_—Å–æ–≤–ø–∞–¥–µ–Ω–∏—è, –°–æ–≤–ø–∞–¥–µ–Ω–∏–µ_%.

    –ò—Å–ø–æ–ª—å–∑—É–µ—Ç fuzzy-–º–∞—Ç—á–∏–Ω–≥ —á–µ—Ä–µ–∑ match_names() –≤–º–µ—Å—Ç–æ —Ç–æ—á–Ω–æ–≥–æ —Å–æ–≤–ø–∞–¥–µ–Ω–∏—è.
    –ü–µ—Ä–µ–∑–∞–ø–∏—Å—ã–≤–∞–µ—Ç —Ç–æ–ª—å–∫–æ –ª–∏—Å—Ç '–ö–ª–∏–µ–Ω—Ç—ã', –æ—Å—Ç–∞–ª—å–Ω—ã–µ –ª–∏—Å—Ç—ã —Å–æ—Ö—Ä–∞–Ω—è—é—Ç—Å—è —á–µ—Ä–µ–∑ openpyxl.
    """
    if not os.path.exists(clients_path):
        log.warning(f"  ‚ö† enrich_clients: —Ñ–∞–π–ª –Ω–µ –Ω–∞–π–¥–µ–Ω: {clients_path}")
        return
    if verification_df is None or len(verification_df) == 0:
        log.warning("  ‚ö† enrich_clients: verification_df –ø—É—Å—Ç–æ–π")
        return

    try:
        from verify_with_db import match_names, normalize_name, normalize_phone
        from config import DB_MATCH_THRESHOLD
    except ImportError:
        log.warning("  ‚ö† enrich_clients: –Ω–µ —É–¥–∞–ª–æ—Å—å –∏–º–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞—Ç—å verify_with_db/config")
        return

    # –ì–æ—Ç–æ–≤–∏–º —Å–ø–∏—Å–æ–∫ –∑–∞–ø–∏—Å–µ–π verification_df –¥–ª—è –ø–µ—Ä–µ–±–æ—Ä–∞
    vdf = verification_df.copy()
    vdf_records = []
    for _, vrow in vdf.iterrows():
        ocr_fio = str(vrow.get("OCR_–§–ò–û", ""))
        ocr_phone = normalize_phone(str(vrow.get("OCR_–¢–µ–ª–µ—Ñ–æ–Ω", "")))
        vdf_records.append({
            "ocr_fio": ocr_fio,
            "ocr_phone": ocr_phone,
            "bd_id": vrow.get("–ë–î_ID", ""),
            "bd_fio": vrow.get("–ë–î_–§–ò–û", ""),
            "status_bd": vrow.get("–°—Ç–∞—Ç—É—Å_–ë–î", ""),
            "score_pct": vrow.get("–°–æ–≤–ø–∞–¥–µ–Ω–∏–µ_%", 0),
        })

    # –ß–∏—Ç–∞–µ–º —Ç–æ–ª—å–∫–æ –ª–∏—Å—Ç '–ö–ª–∏–µ–Ω—Ç—ã' ‚Äî –æ—Å—Ç–∞–ª—å–Ω—ã–µ –Ω–µ —Ç—Ä–æ–≥–∞–µ–º
    from openpyxl import load_workbook
    from zipfile import BadZipFile

    try:
        wb = load_workbook(clients_path)
    except BadZipFile:
        log.warning(f"  ‚ö† enrich_clients: —Ñ–∞–π–ª –ø–æ–≤—Ä–µ–∂–¥—ë–Ω (BadZipFile): {clients_path}")
        return
    except Exception as e:
        log.warning(f"  ‚ö† enrich_clients: –Ω–µ —É–¥–∞–ª–æ—Å—å –æ—Ç–∫—Ä—ã—Ç—å —Ñ–∞–π–ª: {e}")
        return

    if "–ö–ª–∏–µ–Ω—Ç—ã" not in wb.sheetnames:
        log.warning("  ‚ö† enrich_clients: –ª–∏—Å—Ç '–ö–ª–∏–µ–Ω—Ç—ã' –Ω–µ –Ω–∞–π–¥–µ–Ω")
        wb.close()
        return

    try:
        cdf = pd.read_excel(clients_path, sheet_name="–ö–ª–∏–µ–Ω—Ç—ã")
    except Exception as e:
        log.warning(f"  ‚ö† enrich_clients: –Ω–µ —É–¥–∞–ª–æ—Å—å –ø—Ä–æ—á–∏—Ç–∞—Ç—å –ª–∏—Å—Ç –ö–ª–∏–µ–Ω—Ç—ã: {e}")
        wb.close()
        return

    matched_bd_id = []
    matched_bd_fio = []
    matched_status = []
    matched_score = []

    for _, row in cdf.iterrows():
        client_fio = str(row.get("–§–ò–û", ""))
        client_phone = normalize_phone(str(row.get("–¢–µ–ª–µ—Ñ–æ–Ω", "")))

        best_bd_id = ""
        best_bd_fio = ""
        best_status = ""
        best_score = 0.0

        for vrec in vdf_records:
            # –¢–µ–ª–µ—Ñ–æ–Ω –¥–∞—ë—Ç —Ç–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ ‚Äî –º–∞–∫—Å–∏–º–∞–ª—å–Ω—ã–π –ø—Ä–∏–æ—Ä–∏—Ç–µ—Ç
            phone_hit = (client_phone and vrec["ocr_phone"]
                         and client_phone == vrec["ocr_phone"])

            # Fuzzy-–º–∞—Ç—á–∏–Ω–≥ –§–ò–û
            name_score = match_names(client_fio, vrec["ocr_fio"])

            if phone_hit:
                name_score = max(name_score, 0.95)

            if name_score > best_score and name_score >= DB_MATCH_THRESHOLD:
                best_score = name_score
                best_bd_id = vrec["bd_id"]
                best_bd_fio = vrec["bd_fio"]
                best_status = vrec["status_bd"]

        matched_bd_id.append(best_bd_id)
        matched_bd_fio.append(best_bd_fio)
        matched_status.append(best_status)
        matched_score.append(round(best_score * 100, 1) if best_score > 0 else "")

    cdf["–ë–î_ID"] = matched_bd_id
    cdf["–ë–î_–§–ò–û_—Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ"] = matched_bd_fio
    cdf["–°—Ç–∞—Ç—É—Å_—Å–æ–≤–ø–∞–¥–µ–Ω–∏—è"] = matched_status
    cdf["–°–æ–≤–ø–∞–¥–µ–Ω–∏–µ_%"] = matched_score

    # –ü–µ—Ä–µ–∑–∞–ø–∏—Å—ã–≤–∞–µ–º —Ç–æ–ª—å–∫–æ –ª–∏—Å—Ç '–ö–ª–∏–µ–Ω—Ç—ã', —Å–æ—Ö—Ä–∞–Ω—è—è –æ—Å—Ç–∞–ª—å–Ω—ã–µ
    # –£–¥–∞–ª—è–µ–º —Å—Ç–∞—Ä—ã–π –ª–∏—Å—Ç –∏ —Å–æ–∑–¥–∞—ë–º –Ω–æ–≤—ã–π —Å –¥–∞–Ω–Ω—ã–º–∏
    del wb["–ö–ª–∏–µ–Ω—Ç—ã"]
    ws = wb.create_sheet("–ö–ª–∏–µ–Ω—Ç—ã", 0)

    # –ó–∞–≥–æ–ª–æ–≤–∫–∏
    for col_idx, col_name in enumerate(cdf.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # –î–∞–Ω–Ω—ã–µ
    for row_idx, (_, data_row) in enumerate(cdf.iterrows(), 2):
        for col_idx, col_name in enumerate(cdf.columns, 1):
            val = data_row[col_name]
            if pd.isna(val):
                val = ""
            ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(clients_path)
    wb.close()
    log.info(f"  ‚úì –ö–ª–∏–µ–Ω—Ç—ã –¥–æ–ø–æ–ª–Ω–µ–Ω—ã –§–ò–û –∏–∑ –ë–î (fuzzy): {clients_path}")


def main():
    args = parse_args()
    cfg = check_config()

    # –õ–æ–≥–∏—Ä–æ–≤–∞–Ω–∏–µ
    log = setup_pipeline_logging(cfg)

    log.info("")
    log.info("‚ïî‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïó")
    log.info("‚ïë     –ï–î–ò–ù–´–ô –ü–ê–ô–ü–õ–ê–ô–ù –û–¶–ò–§–†–û–í–ö–ò –ö–õ–ò–ï–ù–¢–°–ö–ò–• –ö–ê–†–¢–û–ß–ï–ö  ‚ïë")
    log.info("‚ïë     Google Vision + Claude API + –ë–î –ü—Ä–∏–≤–∏–ª–µ–≥–∏—è      ‚ïë")
    log.info("‚ïö‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïê‚ïù")
    log.info(f"  –í—Ä–µ–º—è –∑–∞–ø—É—Å–∫–∞: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")

    if args.force:
        log.info("  –†–µ–∂–∏–º: --force (–ø–æ–ª–Ω–∞—è –ø–µ—Ä–µ–æ–±—Ä–∞–±–æ—Ç–∫–∞)")
    elif args.skip_ocr:
        log.info("  –†–µ–∂–∏–º: --skip-ocr (—Ç–æ–ª—å–∫–æ —Å–≤–µ—Ä–∫–∞)")
    elif args.only_ocr:
        log.info("  –†–µ–∂–∏–º: --only-ocr (—Ç–æ–ª—å–∫–æ OCR)")
    else:
        log.info("  –†–µ–∂–∏–º: –∏–Ω–∫—Ä–µ–º–µ–Ω—Ç–∞–ª—å–Ω—ã–π (—Ç–æ–ª—å–∫–æ –Ω–æ–≤—ã–µ –∫–∞—Ä—Ç–æ—á–∫–∏)")

    t_start = time.time()
    ocr_excel_path = cfg.OUTPUT_FILE
    normalized_path = None
    verification_df = None

    # ‚îÄ‚îÄ –°–±—Ä–æ—Å —Ä–µ–µ—Å—Ç—Ä–∞ –∏ –∫—ç—à–∞ –ø—Ä–∏ --force ‚îÄ‚îÄ
    if args.force:
        # –°–±—Ä–æ—Å —Ä–µ–µ—Å—Ç—Ä–∞
        reg_path = getattr(cfg, 'PROCESSED_REGISTRY', None)
        if not reg_path:
            reg_path = os.path.join(
                getattr(cfg, 'CACHE_FOLDER', './ocr_cache'),
                "processed_registry.json"
            )
        if os.path.exists(reg_path):
            try:
                os.remove(reg_path)
                log.info(f"  –†–µ–µ—Å—Ç—Ä —Å–±—Ä–æ—à–µ–Ω: {reg_path}")
            except OSError:
                log.warning(f"  –ù–µ —É–¥–∞–ª–æ—Å—å —É–¥–∞–ª–∏—Ç—å —Ä–µ–µ—Å—Ç—Ä: {reg_path}")
        else:
            log.info("  –†–µ–µ—Å—Ç—Ä –Ω–µ –Ω–∞–π–¥–µ–Ω (–∏ —Ç–∞–∫ –ø—É—Å—Ç–æ)")

        # –û—á–∏—Å—Ç–∫–∞ –∫—ç—à–∞ OCR (–≤—Å–µ .json —Ñ–∞–π–ª—ã –±–µ–∑ –∏—Å–∫–ª—é—á–µ–Ω–∏–π)
        cache_folder = getattr(cfg, 'CACHE_FOLDER', './ocr_cache')
        if os.path.exists(cache_folder):
            import glob
            cache_files = glob.glob(os.path.join(cache_folder, "*.json"))
            removed_count = 0
            for cache_file in cache_files:
                try:
                    os.remove(cache_file)
                    removed_count += 1
                except OSError:
                    pass
            if removed_count > 0:
                log.info(f"  –û—á–∏—â–µ–Ω–æ —Ñ–∞–π–ª–æ–≤ –∫—ç—à–∞: {removed_count}")
            else:
                log.info("  –ö—ç—à —É–∂–µ –ø—É—Å—Ç–æ–π")

        # –£–¥–∞–ª—è–µ–º Excel –¥–ª—è –ø–æ–ª–Ω–æ–π –ø–µ—Ä–µ—Å–±–æ—Ä–∫–∏ (–∏–Ω–∞—á–µ –¥–æ–∑–∞–ø–∏—Å—å —Å–æ—Ö—Ä–∞–Ω–∏—Ç —Å—Ç–∞—Ä—ã–µ –¥–∞–Ω–Ω—ã–µ)
        if os.path.exists(ocr_excel_path):
            try:
                os.remove(ocr_excel_path)
                log.info(f"  Excel —Å–±—Ä–æ—à–µ–Ω: {ocr_excel_path}")
            except OSError:
                log.warning(f"  –ù–µ —É–¥–∞–ª–æ—Å—å —É–¥–∞–ª–∏—Ç—å Excel: {ocr_excel_path}")

        # –£–¥–∞–ª—è–µ–º –ø—Ä–æ–º–µ–∂—É—Ç–æ—á–Ω—ã–µ –æ—Ç—á—ë—Ç—ã –¥–ª—è –ø–æ–ª–Ω–æ–π –ø–µ—Ä–µ—Å–±–æ—Ä–∫–∏
        script_dir = os.path.dirname(os.path.abspath(__file__))
        intermediate_files = [
            getattr(cfg, 'NORMALIZED_FILE', 'clients_normalized.xlsx'),
            "verification_report.xlsx",
            "pipeline_report.xlsx",
            getattr(cfg, 'NOT_FOUND_CLIENTS_FILE', 'clients_not_found.xlsx'),
            getattr(cfg, 'FINAL_VERIFICATION_REPORT', 'final_verification_report.xlsx'),
            "raw_results.json",
        ]
        for fname in intermediate_files:
            fpath = os.path.join(script_dir, fname)
            if os.path.exists(fpath):
                try:
                    os.remove(fpath)
                    log.info(f"  –£–¥–∞–ª—ë–Ω: {fname}")
                except OSError:
                    log.warning(f"  –ù–µ —É–¥–∞–ª–æ—Å—å —É–¥–∞–ª–∏—Ç—å: {fname}")

    # ‚îÄ‚îÄ –®–ê–ì 1-4: OCR ‚îÄ‚îÄ
    if not args.skip_ocr:
        check_dependencies()
        ocr_result = run_ocr_pipeline(log, cfg)
        if ocr_result:
            ocr_excel_path = ocr_result
        else:
            log.error("\n  ‚úó OCR –ø–∞–π–ø–ª–∞–π–Ω –∑–∞–≤–µ—Ä—à–∏–ª—Å—è —Å –æ—à–∏–±–∫–æ–π!")
            if not args.only_ocr:
                log.info("  –ü—Ä–æ–±—É—é –ø—Ä–æ–¥–æ–ª–∂–∏—Ç—å —Å–≤–µ—Ä–∫—É —Å —Å—É—â–µ—Å—Ç–≤—É—é—â–∏–º —Ñ–∞–π–ª–æ–º...")
    else:
        if os.path.exists(ocr_excel_path):
            log.info(f"\n  –ò—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è —Å—É—â–µ—Å—Ç–≤—É—é—â–∏–π —Ñ–∞–π–ª: {ocr_excel_path}")
        else:
            log.warning(f"\n  ‚ö† –§–∞–π–ª –Ω–µ –Ω–∞–π–¥–µ–Ω: {ocr_excel_path}")
            log.warning("  –°–≤–µ—Ä–∫–∞ –±—É–¥–µ—Ç —Ç–æ–ª—å–∫–æ –ø–æ –¥–∞–Ω–Ω—ã–º –ë–î.")

    # ‚îÄ‚îÄ –®–ê–ì 4.5: –ù–æ—Ä–º–∞–ª–∏–∑–∞—Ü–∏—è OCR ‚Üí —Ñ–æ—Ä–º–∞—Ç –ë–î ‚îÄ‚îÄ
    if not args.only_ocr:
        normalized_path = run_normalization(log, cfg, ocr_excel_path)

    # ‚îÄ‚îÄ –®–ê–ì 5-6: –°–≤–µ—Ä–∫–∞ ‚îÄ‚îÄ
    # –ò—Å–ø–æ–ª—å–∑—É–µ–º –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω—ã–π —Ñ–∞–π–ª –¥–ª—è —Å–≤–µ—Ä–∫–∏ (–µ—Å–ª–∏ –µ—Å—Ç—å), –∏–Ω–∞—á–µ –æ—Ä–∏–≥–∏–Ω–∞–ª
    if not args.only_ocr:
        verify_path = normalized_path if normalized_path else ocr_excel_path
        report_path, verification_df = run_verification(
            log, cfg, verify_path
        )

        # –ò—Ç–æ–≥–æ–≤—ã–π –∫–æ–º–±–∏–Ω–∏—Ä–æ–≤–∞–Ω–Ω—ã–π –æ—Ç—á—ë—Ç
        generate_pipeline_report(log, cfg, verification_df, ocr_excel_path)

    # ‚îÄ‚îÄ –í—ã–≥—Ä—É–∑–∫–∞ –≤ Google Sheets (–µ—Å–ª–∏ –≤–∫–ª—é—á–µ–Ω–æ) ‚îÄ‚îÄ
    try:
        if _gsheets_disabled(cfg):
            # –í smoke-—Ä–µ–∂–∏–º–µ: —Ç–∏—Ö–∏–π –ø—Ä–æ–ø—É—Å–∫ (–Ω–µ—Ç –ª–∏—à–Ω–µ–≥–æ —à—É–º–∞ –≤ –ª–æ–≥–µ)
            if not _is_smoke_mode():
                log.warning("  ‚ö† –í—ã–≥—Ä—É–∑–∫–∞ –≤ Google Sheets –≤—ã–∫–ª—é—á–µ–Ω–∞ (GSHEETS_UPLOAD_ENABLED=False)")
        else:
            from importlib import import_module
            try:
                google_sheets = import_module('google_sheets')
            except ImportError as e:
                log.warning(f"  ‚ö† Google Sheets –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω (–Ω–µ—Ç –∑–∞–≤–∏—Å–∏–º–æ—Å—Ç–µ–π): {e}")
                google_sheets = None

            creds_path = getattr(cfg, 'GSHEETS_CREDENTIALS', '')
            spreadsheet_id = getattr(cfg, 'GSHEETS_SPREADSHEET_ID', '')
            if google_sheets and creds_path and spreadsheet_id:
                try:
                    if verification_df is not None:
                        google_sheets.upload_df(verification_df, spreadsheet_id, 'verification', creds_path)
                    if os.path.exists(cfg.OUTPUT_FILE):
                        clients_df = pd.read_excel(cfg.OUTPUT_FILE, sheet_name='–ö–ª–∏–µ–Ω—Ç—ã')
                        google_sheets.upload_df(clients_df, spreadsheet_id, 'clients', creds_path)
                    log.info("  ‚úì –í—ã–≥—Ä—É–∂–µ–Ω–æ –≤ Google Sheets")
                except Exception as e:
                    log.warning(f"  ‚ö† –û—à–∏–±–∫–∞ –≤—ã–≥—Ä—É–∑–∫–∏ –≤ Google Sheets: {e}")
            else:
                log.warning("  ‚ö† Google Sheets: –Ω–µ –∑–∞–¥–∞–Ω—ã GSHEETS_CREDENTIALS –∏–ª–∏ GSHEETS_SPREADSHEET_ID")
    except Exception as e:
        log.warning(f"  ‚ö† –û—à–∏–±–∫–∞ –≤ –±–ª–æ–∫–µ –≤—ã–≥—Ä—É–∑–∫–∏ Google Sheets: {e}")

    # ‚îÄ‚îÄ –î–æ–±–∞–≤–ª—è–µ–º –ª–∏—Å—Ç ¬´–°–≤–µ—Ä–∫–∞_–ë–î¬ª –≤ clients_database.xlsx ‚îÄ‚îÄ
    try:
        if verification_df is not None:
            add_verification_sheet(cfg.OUTPUT_FILE, verification_df, log)
    except Exception as e:
        log.warning(f"  ‚ö† –ù–µ —É–¥–∞–ª–æ—Å—å –¥–æ–±–∞–≤–∏—Ç—å –ª–∏—Å—Ç –°–≤–µ—Ä–∫–∞_–ë–î: {e}")

    # ‚îÄ‚îÄ –û–±–æ–≥–∞—â–∞–µ–º clients_database.xlsx –¥–∞–Ω–Ω—ã–º–∏ –ë–î (–§–ò–û –∏–∑ —Å–≤–µ—Ä–∫–∏) ‚îÄ‚îÄ
    try:
        if verification_df is not None:
            enrich_clients_with_db_match(cfg.OUTPUT_FILE, verification_df, log)
    except Exception as e:
        log.warning(f"  ‚ö† –ù–µ —É–¥–∞–ª–æ—Å—å –¥–æ–ø–æ–ª–Ω–∏—Ç—å clients_database.xlsx –§–ò–û –∏–∑ –ë–î: {e}")

    # ‚îÄ‚îÄ –§–∏–Ω–∞–ª—å–Ω–∞—è —Å–≤–æ–¥–∫–∞ ‚îÄ‚îÄ
    total_time = time.time() - t_start
    print_summary(log, verification_df, total_time, ocr_excel_path, cfg)

    log.info("  ‚úì –ü–ê–ô–ü–õ–ê–ô–ù –ó–ê–í–ï–†–®–Å–ù –£–°–ü–ï–®–ù–û")
    log.info("")


if __name__ == "__main__":
    main()
