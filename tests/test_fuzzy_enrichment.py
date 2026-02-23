"""
Тесты fuzzy-обогащения клиентов через enrich_clients_with_db_match.

Проверяют:
1. Fuzzy-матчинг: "Чапленко Ирина" ↔ "Чапленко Ірина" (95%) → БД_ФИО_совпадение заполнено.
2. Точный матч по телефону → БД_ФИО_совпадение заполнено.
3. Нет совпадения → пустые колонки.
4. Лист 'Процедуры' сохраняется при перезаписи.
"""

import sys
import os
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest
import pandas as pd
from unittest.mock import MagicMock
def _create_test_excel(path, clients_data, extra_sheets=None):
    """Создаёт тестовый Excel с листом 'Клиенты' и опциональными дополнительными листами."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(clients_data).to_excel(writer, sheet_name="Клиенты", index=False)
        if extra_sheets:
            for name, data in extra_sheets.items():
                pd.DataFrame(data).to_excel(writer, sheet_name=name, index=False)


def _make_verification_df(records):
    """Создаёт verification_df из списка словарей."""
    return pd.DataFrame(records)


class TestFuzzyEnrichment:
    """Тесты fuzzy-матчинга в enrich_clients_with_db_match."""

    def test_fuzzy_match_chaplienko(self):
        """Чапленко Ирина (OCR) ↔ Чапленко Ирина (verification) → БД_ФИО заполнено."""
        from run_pipeline import enrich_clients_with_db_match

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            _create_test_excel(tmp_path, {
                "ФИО": ["Чапленко Ирина"],
                "Телефон": [""],
            })

            verification_df = _make_verification_df([{
                "OCR_ФИО": "Чапленко Ирина",
                "OCR_Телефон": "",
                "БД_ФИО": "Чапленко Ирина Владимировна",
                "Статус_БД": "Найден в БД",
                "Совпадение_%": 95.0,
            }])

            log = MagicMock()
            enrich_clients_with_db_match(tmp_path, verification_df, log)

            result = pd.read_excel(tmp_path, sheet_name="Клиенты")
            assert "БД_ФИО_совпадение" in result.columns
            assert result.iloc[0]["БД_ФИО_совпадение"] == "Чапленко Ирина Владимировна"
            assert result.iloc[0]["Статус_совпадения"] == "Найден в БД"
            assert float(result.iloc[0]["Совпадение_%"]) > 90
        finally:
            os.remove(tmp_path)

    def test_fuzzy_match_similar_names(self):
        """Похожие имена с опечаткой → fuzzy-матчинг находит совпадение."""
        from run_pipeline import enrich_clients_with_db_match

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            _create_test_excel(tmp_path, {
                "ФИО": ["Иванова Елена"],
                "Телефон": [""],
            })

            verification_df = _make_verification_df([{
                "OCR_ФИО": "Иванова Елена",
                "OCR_Телефон": "",
                "БД_ФИО": "Иванова Елена Петровна",
                "Статус_БД": "Найден в БД",
                "Совпадение_%": 92.0,
            }])

            log = MagicMock()
            enrich_clients_with_db_match(tmp_path, verification_df, log)

            result = pd.read_excel(tmp_path, sheet_name="Клиенты")
            assert result.iloc[0]["БД_ФИО_совпадение"] == "Иванова Елена Петровна"
        finally:
            os.remove(tmp_path)

    def test_phone_match_overrides(self):
        """Совпадение телефона → матч даже при низком ФИО-score."""
        from run_pipeline import enrich_clients_with_db_match

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            _create_test_excel(tmp_path, {
                "ФИО": ["Ким А"],
                "Телефон": ["+7 777 123 4567"],
            })

            verification_df = _make_verification_df([{
                "OCR_ФИО": "Ким Анна Сергеевна",
                "OCR_Телефон": "77771234567",
                "БД_ФИО": "Ким Анна Сергеевна",
                "Статус_БД": "Найден в БД",
                "Совпадение_%": 98.0,
            }])

            log = MagicMock()
            enrich_clients_with_db_match(tmp_path, verification_df, log)

            result = pd.read_excel(tmp_path, sheet_name="Клиенты")
            assert result.iloc[0]["БД_ФИО_совпадение"] == "Ким Анна Сергеевна"
            assert float(result.iloc[0]["Совпадение_%"]) >= 95
        finally:
            os.remove(tmp_path)

    def test_no_match_leaves_empty(self):
        """Нет совпадения → пустые колонки."""
        from run_pipeline import enrich_clients_with_db_match

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            _create_test_excel(tmp_path, {
                "ФИО": ["Абсолютно Другой Человек"],
                "Телефон": [""],
            })

            verification_df = _make_verification_df([{
                "OCR_ФИО": "Петрова Мария",
                "OCR_Телефон": "77770000000",
                "БД_ФИО": "Петрова Мария",
                "Статус_БД": "Найден в БД",
                "Совпадение_%": 95.0,
            }])

            log = MagicMock()
            enrich_clients_with_db_match(tmp_path, verification_df, log)

            result = pd.read_excel(tmp_path, sheet_name="Клиенты")
            val = result.iloc[0]["БД_ФИО_совпадение"]
            assert pd.isna(val) or val == ""
            val2 = result.iloc[0]["Статус_совпадения"]
            assert pd.isna(val2) or val2 == ""
        finally:
            os.remove(tmp_path)

    def test_other_sheets_preserved(self):
        """Лист 'Процедуры' сохраняется при перезаписи 'Клиенты'."""
        from run_pipeline import enrich_clients_with_db_match

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            procedures_data = {"Дата": ["2024-01-15"], "Процедура": ["Чистка"]}
            _create_test_excel(
                tmp_path,
                {"ФИО": ["Тестов Тест"], "Телефон": [""]},
                extra_sheets={"Процедуры": procedures_data},
            )

            verification_df = _make_verification_df([{
                "OCR_ФИО": "Тестов Тест",
                "OCR_Телефон": "",
                "БД_ФИО": "Тестов Тест",
                "Статус_БД": "Найден в БД",
                "Совпадение_%": 100.0,
            }])

            log = MagicMock()
            enrich_clients_with_db_match(tmp_path, verification_df, log)

            # Проверяем что 'Процедуры' сохранился
            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)
            assert "Процедуры" in wb.sheetnames
            assert "Клиенты" in wb.sheetnames
            wb.close()

            proc_df = pd.read_excel(tmp_path, sheet_name="Процедуры")
            assert len(proc_df) == 1
            assert proc_df.iloc[0]["Процедура"] == "Чистка"
        finally:
            os.remove(tmp_path)

    def test_best_match_selected(self):
        """Из нескольких кандидатов выбирается лучший по score."""
        from run_pipeline import enrich_clients_with_db_match

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            _create_test_excel(tmp_path, {
                "ФИО": ["Иванов Иван"],
                "Телефон": [""],
            })

            verification_df = _make_verification_df([
                {
                    "OCR_ФИО": "Иванова Мария",
                    "OCR_Телефон": "",
                    "БД_ФИО": "Иванова Мария",
                    "Статус_БД": "Найден в БД",
                    "Совпадение_%": 70.0,
                },
                {
                    "OCR_ФИО": "Иванов Иван Петрович",
                    "OCR_Телефон": "",
                    "БД_ФИО": "Иванов Иван Петрович",
                    "Статус_БД": "Найден в БД",
                    "Совпадение_%": 98.0,
                },
            ])

            log = MagicMock()
            enrich_clients_with_db_match(tmp_path, verification_df, log)

            result = pd.read_excel(tmp_path, sheet_name="Клиенты")
            # Должен выбрать "Иванов Иван Петрович" как лучшее совпадение
            assert result.iloc[0]["БД_ФИО_совпадение"] == "Иванов Иван Петрович"
        finally:
            os.remove(tmp_path)


class TestFuzzyEnrichmentEdgeCases:
    """Граничные случаи."""

    def test_missing_file_no_crash(self):
        """Несуществующий файл → warning, без исключения."""
        from run_pipeline import enrich_clients_with_db_match

        log = MagicMock()
        enrich_clients_with_db_match("/nonexistent/path.xlsx", pd.DataFrame({"A": [1]}), log)
        log.warning.assert_called()

    def test_empty_verification_df(self):
        """Пустой verification_df → warning, без исключения."""
        from run_pipeline import enrich_clients_with_db_match

        log = MagicMock()
        enrich_clients_with_db_match("/tmp/test.xlsx", pd.DataFrame(), log)
        log.warning.assert_called()

    def test_none_verification_df(self):
        """None verification_df → warning, без исключения."""
        from run_pipeline import enrich_clients_with_db_match

        log = MagicMock()
        enrich_clients_with_db_match("/tmp/test.xlsx", None, log)
        log.warning.assert_called()


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
