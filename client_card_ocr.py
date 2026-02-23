#!/usr/bin/env python3
"""
Скрипт оцифровки бумажных карточек клиентов косметологической клиники.
Google Vision API (OCR) + Claude API (интеллектуальный парсинг) → Excel.

Особенности:
- Нечёткое сопоставление имён (fuzzy matching) для группировки
- Дополнительная привязка по ИИН и телефону
- Кэш для возобновления обработки
- 6 листов Excel с нормализованной структурой
"""

import os
import sys

# Фикс SSL для gRPC на macOS: указываем путь к корневым сертификатам
if not os.environ.get("GRPC_DEFAULT_SSL_ROOTS_FILE_PATH"):
    try:
        import certifi
        os.environ["GRPC_DEFAULT_SSL_ROOTS_FILE_PATH"] = certifi.where()
    except ImportError:
        pass

import json
import time
import base64
import hashlib
import io
import logging
from pathlib import Path
from datetime import datetime
from difflib import SequenceMatcher

import config


# ============================================================
# LAZY IMPORTS — тяжёлые SDK загружаются только при вызове
# ============================================================

def _lazy_import_vision():
    from google.cloud import vision
    return vision


def _lazy_import_service_account():
    from google.oauth2 import service_account
    return service_account


def _lazy_import_anthropic():
    import anthropic
    return anthropic


def _lazy_import_pil_image():
    from PIL import Image
    return Image


_rapidfuzz_cache = {"loaded": False, "fuzz": None}


def _lazy_import_rapidfuzz():
    if not _rapidfuzz_cache["loaded"]:
        try:
            from rapidfuzz import fuzz as rf_fuzz
            _rapidfuzz_cache["fuzz"] = rf_fuzz
        except ImportError:
            _rapidfuzz_cache["fuzz"] = None
        _rapidfuzz_cache["loaded"] = True
    return _rapidfuzz_cache["fuzz"]


_excel_styles_cache = {}


def _get_excel_styles():
    """Lazy-инициализация openpyxl-стилей при первом обращении."""
    if not _excel_styles_cache:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        _excel_styles_cache.update({
            'HEADER_FONT': Font(name='Arial', bold=True, size=11, color='FFFFFF'),
            'HEADER_FILL': PatternFill('solid', fgColor='2F5496'),
            'HEADER_ALIGNMENT': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'CELL_FONT': Font(name='Arial', size=10),
            'CELL_ALIGNMENT': Alignment(vertical='top', wrap_text=True),
            'THIN_BORDER': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'),
            ),
            'WARN_FILL': PatternFill('solid', fgColor='FFF2CC'),
        })
    return _excel_styles_cache


# ============================================================
# 0. НАСТРОЙКА ЛОГИРОВАНИЯ
#    Ежедневные файлы: ocr_logs/ocr_2026-02-09.log
#    Консоль: краткий вывод | Файл: полный вывод с OCR-текстом
# ============================================================

def setup_logging():
    log_folder = getattr(config, 'LOG_FOLDER', './ocr_logs')
    os.makedirs(log_folder, exist_ok=True)

    today = datetime.now().strftime('%Y-%m-%d')
    log_file = os.path.join(log_folder, f"ocr_{today}.log")

    level_str = getattr(config, 'LOG_LEVEL', 'DEBUG').upper()
    level = getattr(logging, level_str, logging.DEBUG)

    logger = logging.getLogger('ocr')
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    # Файл — полный лог (DEBUG и выше)
    fh = logging.FileHandler(log_file, encoding='utf-8')
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter(
        '%(asctime)s | %(levelname)-7s | %(message)s',
        datefmt='%H:%M:%S'
    ))
    logger.addHandler(fh)

    # Консоль — краткий вывод (INFO и выше)
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(level)
    ch.setFormatter(logging.Formatter('%(message)s'))
    logger.addHandler(ch)

    logger.info(f"Лог-файл: {log_file}")
    return logger


log = setup_logging()


# ============================================================
# 1. ИНИЦИАЛИЗАЦИЯ
# ============================================================

def init_vision_client():
    service_account = _lazy_import_service_account()
    vision = _lazy_import_vision()
    credentials = service_account.Credentials.from_service_account_file(
        config.GOOGLE_VISION_CREDENTIALS
    )
    return vision.ImageAnnotatorClient(credentials=credentials)


def init_claude_client():
    anthropic = _lazy_import_anthropic()
    base_url = getattr(config, 'ANTHROPIC_BASE_URL', None)
    if base_url:
        return anthropic.Anthropic(api_key=config.ANTHROPIC_API_KEY, base_url=base_url)
    return anthropic.Anthropic(api_key=config.ANTHROPIC_API_KEY)


# ============================================================
# 2. GOOGLE VISION OCR
# ============================================================

def ocr_image_structured(vision_client, image_path: str):
    """
    Вызывает Vision API и возвращает структурированный результат
    включая блоки, реконструированные таблицы и обогащённый текст.

    Если OCR_TABLE_RECONSTRUCTION и OCR_SAVE_BBOX_DEBUG оба выключены —
    структурные блоки не извлекаются (экономия времени).

    Returns:
        OcrStructuredResult
    """
    from utils.table_reconstruction import (
        extract_structured_blocks, reconstruct_all_tables,
        build_enhanced_text, save_bbox_debug, OcrStructuredResult,
    )
    vision = _lazy_import_vision()

    filename = os.path.basename(image_path)
    file_size = os.path.getsize(image_path) / 1024  # KB
    log.debug(f"[OCR] Отправка в Vision API: {filename} ({file_size:.0f} KB)")

    with open(image_path, 'rb') as f:
        content = f.read()

    image = vision.Image(content=content)
    t0 = time.time()
    response = vision_client.document_text_detection(
        image=image,
        image_context=vision.ImageContext(language_hints=['ru'])
    )
    elapsed = time.time() - t0

    if response.error.message:
        log.error(f"[OCR] Vision API ошибка для {filename}: {response.error.message}")
        raise Exception(f"Vision API ошибка: {response.error.message}")

    full_text = ""
    if response.full_text_annotation:
        full_text = response.full_text_annotation.text
        log.debug(f"[OCR] {filename}: распознано {len(full_text)} символов за {elapsed:.1f}с")
        log.debug(f"[OCR] {filename} текст (первые 200 сим.): {full_text[:200].replace(chr(10), ' ')}")
    else:
        log.warning(f"[OCR] {filename}: текст не найден (пустой ответ)")

    need_tables = getattr(config, 'OCR_TABLE_RECONSTRUCTION', True)
    need_debug = getattr(config, 'OCR_SAVE_BBOX_DEBUG', False)

    # Извлекаем структурные блоки только если нужна реконструкция или debug
    blocks = []
    table_blocks = []
    tables_md = ""
    tables_csv = ""
    page_confidence = 0.0

    if need_tables or need_debug:
        blocks = extract_structured_blocks(response)
        table_blocks = [b for b in blocks if b.block_type == 2]

        if blocks:
            page_confidence = sum(b.confidence for b in blocks) / len(blocks)

        # Реконструкция таблиц
        if need_tables and table_blocks:
            row_tol = getattr(config, 'TABLE_ROW_TOLERANCE_PX', 15)
            tables_md, tables_csv = reconstruct_all_tables(blocks, row_tol)
            if tables_md:
                log.info(f"  [TABLE] {filename}: реконструировано {len(table_blocks)} таблиц ({len(tables_md)} сим.)")

        # Debug: bbox/confidence
        if need_debug:
            debug_folder = getattr(config, 'OCR_DEBUG_FOLDER', './ocr_debug')
            debug_path = save_bbox_debug(image_path, blocks, page_confidence, debug_folder)
            log.debug(f"[DEBUG] bbox/confidence saved: {debug_path}")

    enhanced_text = build_enhanced_text(full_text, tables_md)

    return OcrStructuredResult(
        full_text=full_text,
        blocks=blocks,
        table_blocks=table_blocks,
        tables_md=tables_md,
        tables_csv=tables_csv,
        enhanced_text=enhanced_text,
        page_confidence=page_confidence,
    )


def ocr_image(vision_client, image_path: str) -> str:
    """Обратно совместимая обёртка — возвращает только текст."""
    result = ocr_image_structured(vision_client, image_path)
    return result.full_text


# ============================================================
# 3. CLAUDE API — ИНТЕЛЛЕКТУАЛЬНЫЙ ПАРСИНГ
# ============================================================

def _load_prompt() -> str:
    """Загружает системный промпт из prompts/claude_ocr_prompt.md."""
    prompt_path = os.path.join(os.path.dirname(__file__), "prompts", "claude_ocr_prompt.md")
    try:
        with open(prompt_path, "r", encoding="utf-8") as f:
            return f.read().strip()
    except FileNotFoundError:
        log.warning(f"[PROMPT] Файл промпта не найден: {prompt_path}, используем встроенный fallback")
        return (
            "You are a strict JSON-only OCR extraction system. "
            "Output ONLY a raw JSON object with keys: page_type, data, raw_text. "
            "No markdown, no comments, no explanation."
        )


CLAUDE_SYSTEM_PROMPT = _load_prompt()


def image_to_base64(image_path: str) -> tuple:
    with open(image_path, 'rb') as f:
        data = f.read()

    ext = Path(image_path).suffix.lower()
    media_types = {
        '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
        '.png': 'image/png', '.webp': 'image/webp'
    }
    media_type = media_types.get(ext, 'image/jpeg')
    return base64.standard_b64encode(data).decode('utf-8'), media_type


def _pick_first(data: dict, paths: list):
    """Возвращает первое непустое значение по списку путей.

    Путь может быть строкой ('fio') или кортежем ('patient_info', 'name').
    """
    for path in paths:
        cur = data
        if isinstance(path, str):
            path = path.split('.')
        ok = True
        for key in path:
            if isinstance(cur, dict) and key in cur:
                cur = cur[key]
            else:
                ok = False
                break
        if ok and cur not in (None, "", float('nan')):
            # Пропускаем словари/списки — ищем скалярные значения
            if isinstance(cur, (dict, list)):
                continue
            return cur
    return None


def _normalize_simple(value: str) -> str:
    if value is None:
        return ""
    s = str(value).replace('ё', 'е')
    return " ".join(s.split())


def collect_name_phone_iin(data: dict) -> dict:
    """Обогащает payload каноническими fio/phone/iin из разных алиасов."""
    if not isinstance(data, dict):
        return {"fio": "", "phone": "", "iin": ""}

    aliases_fio = [
        ("пациент", "фио"),
        ("patient_info", "name"),
        "fio",
        "patient_name",
        ("document", "patient_name"),
        "client_name",
        "patient",
    ]
    aliases_phone = [
        "phone", "contact", "contacts",
        ("patient_info", "phone"),
        ("пациент", "телефон"),
    ]
    aliases_iin = [
        ("пациент", "иин"),
        "iin",
        ("patient_info", "iin"),
    ]

    fio = _pick_first(data, aliases_fio)
    phone = _pick_first(data, aliases_phone)
    iin = _pick_first(data, aliases_iin)

    enriched = dict(data)
    if fio:
        enriched.setdefault("fio", _normalize_simple(fio))
    if phone:
        enriched.setdefault("phone", _normalize_simple(phone))
    if iin:
        enriched.setdefault("iin", _normalize_simple(iin))
    return enriched


def normalize_claude_response(payload: dict, ocr_text: str, filename: str) -> dict:
    """
    Нормализует ответ Claude к каноническому формату.

    Поддерживает форматы:
    1. Канонический: {page_type, data}
    2. Альтернативный: {document_type, ...fields...}
    3. Русские ключи в корне: {медицинская_карта: {...}}
    4. JSON в markdown-обертке (обрабатывается выше)

    Возвращает:
        {
            page_type: str,
            data: dict,
            raw_payload: dict,
            parse_mode: 'strict' | 'recovered' | 'fallback'
        }
    """
    if not payload or not isinstance(payload, dict):
        return {
            "page_type": "unknown",
            "data": {},
            "raw_payload": payload,
            "parse_mode": "fallback"
        }

    # Режим 1: Канонический формат (page_type + data)
    if "page_type" in payload and "data" in payload:
        page_type = payload.get("page_type", "unknown")
        data = payload.get("data", {})

        # Валидация page_type
        valid_types = {
            "medical_card_front", "medical_card_inner", "procedure_sheet",
            "products_list", "complex_package", "botox_record", "unknown"
        }
        if page_type not in valid_types:
            log.debug(f"[NORMALIZE] {filename}: некорректный page_type='{page_type}', ставлю unknown")
            page_type = "unknown"

        data = data if isinstance(data, dict) else {}
        data = collect_name_phone_iin(data)
        return {
            "page_type": page_type,
            "data": data,
            "raw_payload": payload,
            "parse_mode": "strict"
        }

    # Режим 2: document_type формат
    if "document_type" in payload:
        doc_type = payload.get("document_type", "").lower()

        # Маппинг document_type → page_type
        doc_type_map = {
            "medical_card_front": "medical_card_front",
            "medical_card_inner": "medical_card_inner",
            "procedure_sheet": "procedure_sheet",
            "products_list": "products_list",
            "complex_package": "complex_package",
            "botox_record": "botox_record",
            "медицинская карта": "medical_card_front",
            "медицинская_карта": "medical_card_front",
            "процедурный лист": "procedure_sheet",
            "процедурный_лист": "procedure_sheet",
            "покупки": "products_list",
            "список приобретенных средств для домашнего ухода": "products_list",
            "список приобретённых средств для домашнего ухода": "products_list",
            "комплекс": "complex_package",
            "ботокс": "botox_record",
            "ботулинический токсин": "botox_record",
        }

        page_type = doc_type_map.get(doc_type, "unknown")

        # Собираем data из остальных полей
        data = {k: v for k, v in payload.items() if k != "document_type"}

        log.debug(f"[NORMALIZE] {filename}: восстановлен из document_type='{doc_type}' → {page_type}")

        data = collect_name_phone_iin(data)
        return {
            "page_type": page_type,
            "data": data,
            "raw_payload": payload,
            "parse_mode": "recovered"
        }

    # Режим 3: Русские ключи в корне
    russian_keys = {
        "медицинская_карта": "medical_card_front",
        "медкарта": "medical_card_front",
        "процедурный_лист": "procedure_sheet",
        "процедуры": "procedure_sheet",
        "покупки": "products_list",
        "косметика": "products_list",
        "комплекс": "complex_package",
        "пакет": "complex_package",
        "ботокс": "botox_record",
    }

    for rus_key, page_type in russian_keys.items():
        if rus_key in payload:
            data = payload.get(rus_key, {})
            if isinstance(data, dict):
                log.debug(f"[NORMALIZE] {filename}: восстановлен из русского ключа '{rus_key}' → {page_type}")
                data = collect_name_phone_iin(data)
                return {
                    "page_type": page_type,
                    "data": data,
                    "raw_payload": payload,
                    "parse_mode": "recovered"
                }

    # Режим 4: Keyword heuristics на основе содержимого payload и OCR-текста
    page_type = infer_page_type_from_content(payload, ocr_text, filename)

    # Если не удалось определить - оставляем unknown
    if page_type == "unknown":
        log.warning(f"[NORMALIZE] {filename}: не удалось определить тип страницы, ключи: {list(payload.keys())[:10]}")

    payload = collect_name_phone_iin(payload)
    return {
        "page_type": page_type,
        "data": payload,  # Весь payload идёт в data
        "raw_payload": payload,
        "parse_mode": "recovered" if page_type != "unknown" else "fallback"
    }


def infer_page_type_from_content(payload: dict, ocr_text: str, filename: str) -> str:
    """
    Определяет тип страницы по ключевым словам в payload и OCR-тексте.

    Используется как fallback когда формат ответа нестандартный.
    """
    # Нормализуем ключи payload и OCR-текст для поиска
    keys_lower = [str(k).lower() for k in payload.keys()]
    text_lower = ocr_text.lower() if ocr_text else ""

    # Характерные признаки типов страниц
    indicators = {
        "medical_card_front": [
            "fio", "фио", "birth_date", "рождение", "iin", "иин",
            "citizenship", "гражданство", "address", "адрес",
            "allergies", "аллергии", "emergency_contact", "экстренный"
        ],
        "medical_card_inner": [
            "complaints", "жалобы", "objective_status", "статус",
            "diagnosis", "диагноз", "blood_pressure", "давление",
            "hepatitis", "гепатит", "chronic", "хронические"
        ],
        "procedure_sheet": [
            "procedures", "процедуры", "procedure_name", "название",
            "description", "описание", "процедур"
        ],
        "products_list": [
            "products", "покупки", "product_name", "товар",
            "consultant", "консультант", "средств", "косметика"
        ],
        "complex_package": [
            "complex_name", "комплекс", "package", "пакет",
            "purchase_date", "приобретения", "privilage", "привилегия"
        ],
        "botox_record": [
            "injections", "инъекции", "drug", "препарат",
            "injection_area", "область", "botox", "ботокс", "ботулин"
        ],
    }

    # Подсчитываем совпадения для каждого типа
    scores = {}
    for page_type, keywords in indicators.items():
        score = 0
        for keyword in keywords:
            # Проверяем в ключах payload
            if any(keyword in key for key in keys_lower):
                score += 2
            # Проверяем в OCR-тексте (меньший вес)
            if keyword in text_lower:
                score += 1
        scores[page_type] = score

    # Выбираем тип с максимальным score
    if scores:
        best_type = max(scores, key=scores.get)
        best_score = scores[best_type]

        # Порог для уверенности: минимум 3 совпадения
        if best_score >= 3:
            log.debug(f"[INFER] {filename}: определён как {best_type} (score={best_score})")
            return best_type

    # Не удалось определить
    return "unknown"


def extract_with_claude(claude_client, image_path: str, ocr_text: str) -> dict:
    filename = os.path.basename(image_path)
    log.debug(f"[CLAUDE] Отправка в Claude: {filename}")

    img_b64, media_type = image_to_base64(image_path)

    # Сжимаем если > 5MB
    if len(img_b64) > 5_000_000:
        log.debug(f"[CLAUDE] {filename}: сжатие изображения (>{len(img_b64)//1024}KB)")
        Image = _lazy_import_pil_image()
        img = Image.open(image_path)
        img.thumbnail((2000, 2000), Image.Resampling.LANCZOS)
        buffer = io.BytesIO()
        img.save(buffer, format='JPEG', quality=85)
        img_b64 = base64.standard_b64encode(buffer.getvalue()).decode('utf-8')
        media_type = 'image/jpeg'

    t0 = time.time()
    message = claude_client.messages.create(
        model=config.CLAUDE_MODEL,
        max_tokens=4096,
        system=CLAUDE_SYSTEM_PROMPT,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": media_type,
                        "data": img_b64
                    }
                },
                {
                    "type": "text",
                    "text": (
                        f"OCR-текст:\n\n{ocr_text}\n\n"
                        "Extract ALL data into JSON. "
                        "Reply with ONLY the raw JSON object — no markdown fences, no explanation."
                    )
                }
            ]
        }]
    )
    elapsed = time.time() - t0

    response_text = message.content[0].text.strip()
    tokens_in = getattr(message.usage, 'input_tokens', '?')
    tokens_out = getattr(message.usage, 'output_tokens', '?')
    log.debug(f"[CLAUDE] {filename}: ответ за {elapsed:.1f}с (токены: {tokens_in}→{tokens_out})")

    # Убираем markdown-обёртки
    if response_text.startswith('```'):
        response_text = response_text.split('\n', 1)[1]
        if response_text.endswith('```'):
            response_text = response_text[:-3]
        response_text = response_text.strip()

    try:
        parsed = json.loads(response_text)
        normalized = normalize_claude_response(parsed, ocr_text, filename)
        log.debug(f"[CLAUDE] {filename}: тип={normalized.get('page_type','?')} (режим: {normalized.get('parse_mode','?')})")
        log.debug(f"[CLAUDE] {filename}: данные={json.dumps(normalized.get('data',{}), ensure_ascii=False)[:500]}")
        return normalized
    except json.JSONDecodeError:
        start = response_text.find('{')
        end = response_text.rfind('}') + 1
        if start >= 0 and end > start:
            try:
                parsed = json.loads(response_text[start:end])
                normalized = normalize_claude_response(parsed, ocr_text, filename)
                log.warning(f"[CLAUDE] {filename}: JSON извлечён из текста (тип={normalized.get('page_type','?')})")
                return normalized
            except json.JSONDecodeError:
                pass
        log.error(f"[CLAUDE] {filename}: НЕ УДАЛОСЬ распарсить JSON. Ответ: {response_text[:300]}")
        return {
            "page_type": "unknown",
            "data": {"raw_text": response_text},
            "raw_payload": response_text[:1000],
            "parse_mode": "fallback"
        }


# ============================================================
# 3.1. КОРРЕКЦИЯ ИМЁН ВРАЧЕЙ
#      Нечёткое сопоставление с известным списком из config.py
# ============================================================

def _build_doctor_aliases():
    """
    Строит словарь сокращений → полное имя врача.
    Например: "О.А." → "Асшеман Оксана", "Виктория" → "Житникова Виктория"
    """
    doctors = getattr(config, 'KNOWN_DOCTORS', [])
    aliases = {}
    for doc in doctors:
        parts = doc.split()
        if len(parts) >= 2:
            surname = parts[0]
            first_name = parts[1]
            # Полное имя
            aliases[doc.lower()] = doc
            aliases[f"{surname} {first_name}".lower()] = doc
            aliases[f"{first_name} {surname}".lower()] = doc
            # Только имя
            aliases[first_name.lower()] = doc
            # Инициалы: "О.А." (Имя.Фамилия)
            initials = f"{first_name[0]}.{surname[0]}."
            aliases[initials.lower()] = doc
            # Обратные инициалы: "А.О."
            initials_rev = f"{surname[0]}.{first_name[0]}."
            aliases[initials_rev.lower()] = doc
            # Сокращённые имена
            short_names = {
                "виктория": ["вика", "виека", "викт"],
                "оксана": ["оксан", "окс"],
                "эльвира": ["эля", "эльв"],
                "ольга": ["оля"],
                "рада": ["рад"],
            }
            for full_short, variants in short_names.items():
                if first_name.lower() == full_short:
                    for v in variants:
                        aliases[v] = doc
    return aliases


DOCTOR_ALIASES = _build_doctor_aliases()


def correct_doctor_name(raw_name: str) -> str:
    """
    Пытается найти точное совпадение с известным врачом.
    Если не нашли — пробуем нечёткое сопоставление (порог 0.6).
    """
    if not raw_name:
        return raw_name

    cleaned = raw_name.strip()

    # 1. Точное совпадение по алиасам
    if cleaned.lower() in DOCTOR_ALIASES:
        matched = DOCTOR_ALIASES[cleaned.lower()]
        if matched != cleaned:
            log.debug(f"[ВРАЧ] «{cleaned}» → «{matched}» (точное совпадение)")
        return matched

    # 2. Нечёткое совпадение с известными врачами
    doctors = getattr(config, 'KNOWN_DOCTORS', [])
    best_match = None
    best_score = 0.0

    for doc in doctors:
        score = fuzzy_match(cleaned, doc)
        if score > best_score:
            best_score = score
            best_match = doc

    if best_score >= 0.6 and best_match:
        log.debug(f"[ВРАЧ] «{cleaned}» → «{best_match}» (fuzzy: {best_score:.0%})")
        return best_match

    log.debug(f"[ВРАЧ] «{cleaned}» — не найден в списке (лучший: {best_match}, {best_score:.0%})")
    return cleaned


def correct_doctors_in_data(data: dict) -> dict:
    """Корректирует имена врачей/консультантов во всех полях результата."""
    # Прямые поля
    for key in ["doctor", "consultant"]:
        if data.get(key):
            data[key] = correct_doctor_name(data[key])

    # В массивах процедур/покупок
    for list_key in ["procedures", "products", "injections"]:
        items = data.get(list_key, [])
        if isinstance(items, list):
            for item in items:
                if isinstance(item, dict):
                    for key in ["doctor", "consultant"]:
                        if item.get(key):
                            item[key] = correct_doctor_name(item[key])

    return data


# ============================================================
# 4. КЭШ
# ============================================================

def get_cache_path(image_path: str) -> str:
    os.makedirs(config.CACHE_FOLDER, exist_ok=True)
    h = hashlib.md5(image_path.encode()).hexdigest()
    return os.path.join(config.CACHE_FOLDER, f"{h}.json")


def load_from_cache(image_path: str) -> dict | None:
    path = get_cache_path(image_path)
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return None


def save_to_cache(image_path: str, result: dict):
    with open(get_cache_path(image_path), 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)


# ============================================================
# 4.1. РЕЕСТР ОБРАБОТАННЫХ КАРТОЧЕК
#      Хранит список файлов, уже записанных в итоговый Excel.
#      При повторном запуске эти файлы ПРОПУСКАЮТСЯ.
#      Удалите файл реестра, чтобы обработать всё заново.
# ============================================================

def _get_registry_path() -> str:
    """Путь к файлу реестра обработанных карточек."""
    reg = getattr(config, 'PROCESSED_REGISTRY', None)
    if reg:
        return reg
    return os.path.join(config.CACHE_FOLDER, "processed_registry.json")


def load_registry() -> dict:
    """
    Загружает реестр обработанных карточек.
    Формат: {filename: {md5, page_type, client_name, processed_at, written_to_excel}}
    """
    path = _get_registry_path()
    if os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            log.warning(f"[РЕЕСТР] Не удалось прочитать {path}, начинаю заново")
    return {}


def save_registry(registry: dict):
    """Сохраняет реестр на диск."""
    path = _get_registry_path()
    os.makedirs(os.path.dirname(path) or '.', exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(registry, f, ensure_ascii=False, indent=2)
    log.debug(f"[РЕЕСТР] Сохранён: {len(registry)} записей → {path}")


def register_processed(registry: dict, result: dict):
    """Добавляет обработанный файл в реестр."""
    filename = result.get("filename", "")
    if not filename:
        return
    fio = (result.get("data", {}).get("fio")
           or result.get("data", {}).get("patient_name")
           or "—")
    registry[filename] = {
        "md5": hashlib.md5(result.get("filepath", "").encode()).hexdigest(),
        "page_type": result.get("page_type", "unknown"),
        "client_name": fio,
        "processed_at": result.get("processed_at", datetime.now().isoformat()),
        "written_to_excel": True,
    }


def _file_md5(filepath: str) -> str:
    """MD5-хэш содержимого файла (для обнаружения изменений)."""
    try:
        h = hashlib.md5()
        with open(filepath, 'rb') as f:
            for chunk in iter(lambda: f.read(8192), b''):
                h.update(chunk)
        return h.hexdigest()
    except (IOError, OSError):
        return ""


# ============================================================
# 5. ОБРАБОТКА ИЗОБРАЖЕНИЙ
# ============================================================

def get_image_files(folder: str) -> list:
    files = []
    for f in sorted(os.listdir(folder)):
        if Path(f).suffix.lower() in config.IMAGE_EXTENSIONS:
            files.append(os.path.join(folder, f))
    return files


def process_all_images(vision_client, claude_client) -> list:
    image_files = get_image_files(config.INPUT_FOLDER)

    if not image_files:
        log.error(f"Фотографии не найдены в: {config.INPUT_FOLDER}")
        sys.exit(1)

    # ── Загружаем реестр обработанных ──
    registry = load_registry()
    already_done = set(registry.keys())

    # Фильтруем: только новые файлы
    new_files = []
    skipped_files = []
    for img_path in image_files:
        filename = os.path.basename(img_path)
        if filename in already_done:
            # Проверяем: файл не изменился? (MD5 содержимого)
            old_md5 = registry[filename].get("md5", "")
            # md5 в реестре — это хэш пути, не содержимого (для совместимости с кэшем)
            # Просто пропускаем
            skipped_files.append(filename)
        else:
            new_files.append(img_path)

    log.info(f"\nНайдено {len(image_files)} фото")
    if skipped_files:
        log.info(f"  ✓ Уже обработано (пропуск): {len(skipped_files)}")
    log.info(f"  → Новых для обработки: {len(new_files)}")

    if not new_files:
        log.info("\n  Все карточки уже обработаны. Новых нет.")
        log.info("  (Удалите реестр для повторной обработки)")
        # Возвращаем результаты из кэша для записи в Excel
        cached_results = []
        for img_path in image_files:
            cached = load_from_cache(img_path)
            if cached:
                cached_results.append(cached)
        return cached_results

    log.info("")
    log.debug(f"Новые файлы: {[os.path.basename(f) for f in new_files]}")

    results = []
    errors = []

    # Сначала загружаем уже обработанные из кэша (для группировки)
    for img_path in image_files:
        filename = os.path.basename(img_path)
        if filename in already_done:
            cached = load_from_cache(img_path)
            if cached:
                results.append(cached)

    # Обрабатываем ТОЛЬКО новые
    from tqdm import tqdm
    for idx, img_path in enumerate(tqdm(new_files, desc="Обработка", unit="фото"), 1):
        filename = os.path.basename(img_path)
        log.info(f"\n{'─'*50}")
        log.info(f"[{idx}/{len(new_files)}] {filename}")

        # Проверяем кэш (на случай если файл в кэше, но не в реестре)
        cached = load_from_cache(img_path)
        if cached:
            log.info(f"  ↩ из кэша (тип: {cached.get('page_type','?')})")
            results.append(cached)
            register_processed(registry, cached)
            continue

        for attempt in range(config.MAX_RETRIES + 1):
            try:
                ocr_result = ocr_image_structured(vision_client, img_path)
                # Claude получает enhanced_text (с таблицами), дедуп/Excel — оригинальный текст
                parsed = extract_with_claude(claude_client, img_path, ocr_result.enhanced_text)

                # Коррекция имён врачей
                if parsed.get("data"):
                    parsed["data"] = correct_doctors_in_data(parsed["data"])

                page_type = parsed.get("page_type", "unknown")
                fio = parsed.get("data", {}).get("fio") or parsed.get("data", {}).get("patient_name") or "—"

                result = {
                    "filename": filename,
                    "filepath": img_path,
                    "ocr_text": ocr_result.full_text,
                    "tables_md": ocr_result.tables_md,
                    "tables_csv": ocr_result.tables_csv,
                    "page_confidence": round(ocr_result.page_confidence, 4),
                    "page_type": page_type,
                    "data": parsed.get("data", {}),
                    "raw_payload": parsed.get("raw_payload"),
                    "parse_mode": parsed.get("parse_mode", "unknown"),
                    "processed_at": datetime.now().isoformat()
                }
                save_to_cache(img_path, result)
                results.append(result)

                # Записываем в реестр
                register_processed(registry, result)

                log.info(f"  ✓ тип: {page_type} | клиент: {fio}")
                time.sleep(config.API_DELAY)
                break

            except Exception as e:
                if attempt < config.MAX_RETRIES:
                    wait = 2 ** (attempt + 1)
                    log.warning(f"  ⚠ Попытка {attempt+1}/{config.MAX_RETRIES}: {e}")
                    log.warning(f"    Ожидание {wait}с перед повтором...")
                    time.sleep(wait)
                else:
                    log.error(f"  ✗ ОТКАЗ после {config.MAX_RETRIES} попыток: {e}")
                    errors.append(f"{filename}: {e}")
                    results.append({
                        "filename": filename, "filepath": img_path,
                        "page_type": "error", "data": {"error": str(e)},
                        "processed_at": datetime.now().isoformat()
                    })

    # Сохраняем реестр после обработки
    save_registry(registry)

    log.info(f"\n{'═'*50}")
    log.info(f"Обработка завершена: {len(new_files)} новых, "
             f"{len(skipped_files)} пропущено, ошибок: {len(errors)}")
    if errors:
        for e in errors:
            log.error(f"  - {e}")

    return results


# ============================================================
# 6. УМНАЯ ГРУППИРОВКА ПО КЛИЕНТАМ
#    (нечёткое сопоставление + привязка по ИИН/телефону)
# ============================================================

def normalize_name(name: str) -> str:
    """Убирает лишние пробелы, нижний регистр, сортирует слова."""
    if not name:
        return ""
    words = name.strip().lower().split()
    return " ".join(sorted(words))  # Сортируем слова, чтобы "Иванов Пётр" = "Пётр Иванов"


def fuzzy_match(name1: str, name2: str) -> float:
    """
    Возвращает степень совпадения двух имён (0.0 - 1.0).
    Использует rapidfuzz если доступен, иначе difflib.
    """
    if not name1 or not name2:
        return 0.0

    n1 = normalize_name(name1)
    n2 = normalize_name(name2)

    if n1 == n2:
        return 1.0

    rf_fuzz = _lazy_import_rapidfuzz()
    if rf_fuzz is not None:
        # token_sort_ratio отлично работает для "Иванов Пётр" vs "Пётр Иванов"
        return rf_fuzz.token_sort_ratio(n1, n2) / 100.0
    else:
        return SequenceMatcher(None, n1, n2).ratio()


def extract_identifiers(result: dict) -> dict:
    """Извлекает все идентификаторы клиента из результата."""
    data = result.get("data", {})
    data = collect_name_phone_iin(data)
    fio = data.get("fio", "")
    phone = data.get("phone", "")
    iin = data.get("iin", "")
    # Нормализация пробелов и ё/е уже в helper
    return {
        "fio": fio,
        "phone": phone,
        "iin": iin,
    }


def find_matching_client(identifiers: dict, clients: dict, threshold: float) -> str | None:
    """
    Ищет существующего клиента по нечёткому совпадению.

    Логика приоритетов:
    1. Точное совпадение ИИН → 100% тот же клиент
    2. Точное совпадение телефона → 100% тот же клиент
    3. Нечёткое совпадение ФИО >= threshold → вероятно тот же клиент
    """
    new_fio = identifiers["fio"]
    new_phone = identifiers["phone"].replace(" ", "").replace("-", "").replace("+", "")
    new_iin = identifiers["iin"].replace(" ", "")

    for client_key, client_data in clients.items():
        # Проверка 1: ИИН (если есть у обоих)
        if new_iin and client_data.get("iin"):
            client_iin = client_data["iin"].replace(" ", "")
            if new_iin == client_iin:
                return client_key

        # Проверка 2: Телефон (если есть у обоих)
        if new_phone and len(new_phone) >= 7 and client_data.get("phone"):
            client_phone = client_data["phone"].replace(" ", "").replace("-", "").replace("+", "")
            if client_phone and new_phone[-7:] == client_phone[-7:]:
                return client_key

        # Проверка 3: Нечёткое ФИО
        if new_fio and client_data.get("name"):
            similarity = fuzzy_match(new_fio, client_data["name"])
            if similarity >= threshold:
                return client_key

    return None


def group_by_client(results: list) -> dict:
    """
    Группирует результаты по клиентам с нечётким сопоставлением.

    Возвращает: {client_key: {name, phone, iin, pages: [...]}}
    """
    clients = {}
    threshold = getattr(config, 'FUZZY_NAME_THRESHOLD', 0.75)
    unmatched = []

    # Сначала обрабатываем medical_card_front — они содержат больше всего идентификаторов
    priority_order = ['medical_card_front', 'complex_package', 'procedure_sheet',
                      'products_list', 'medical_card_inner', 'botox_record']

    sorted_results = sorted(
        results,
        key=lambda r: priority_order.index(r.get("page_type", ""))
            if r.get("page_type", "") in priority_order else 99
    )

    for result in sorted_results:
        if result.get("page_type") == "error":
            unmatched.append(result)
            continue

        ids = extract_identifiers(result)
        if not ids["fio"] and not ids["phone"] and not ids["iin"]:
            unmatched.append(result)
            continue

        # Ищем совпадение с существующим клиентом
        match_key = find_matching_client(ids, clients, threshold)

        if match_key:
            clients[match_key]["pages"].append(result)
            log.debug(f"[ГРУППИРОВКА] «{ids['fio']}» → привязан к «{clients[match_key]['name']}»")
            if ids["phone"] and not clients[match_key].get("phone"):
                clients[match_key]["phone"] = ids["phone"]
            if ids["iin"] and not clients[match_key].get("iin"):
                clients[match_key]["iin"] = ids["iin"]
        else:
            name = ids["fio"] if ids["fio"] else "(без ФИО)"
            client_key = f"client_{len(clients)+1}"
            clients[client_key] = {
                "name": name,
                "phone": ids["phone"],
                "iin": ids["iin"],
                "pages": [result]
            }
            log.debug(f"[ГРУППИРОВКА] Новый клиент: «{name}» → {client_key}")

    # Непривязанные страницы → в отдельную группу (не трогаем в dedup)
    if unmatched:
        # Если клиентов нет вовсе — создаём одного, даже без fio, чтобы тексты не потерять
        existing_keys = [k for k in clients if k != "_unmatched"]
        if len(existing_keys) == 0:
            any_ids = extract_identifiers(unmatched[0])
            name = any_ids["fio"] or ("(без ФИО)" if (any_ids["phone"] or any_ids["iin"]) else "⚠ Не удалось определить клиента")
            clients["client_1"] = {
                "name": name,
                "phone": any_ids["phone"],
                "iin": any_ids["iin"],
                "pages": unmatched,
            }
        else:
            # Пробуем приклеить непривязанные страницы к ближайшему клиенту (по similarity), иначе кладём всё в первого клиента
            primary_key = existing_keys[0]
            remaining = []
            for page in unmatched:
                ids = extract_identifiers(page)
                match_key = None
                if ids["fio"] or ids["phone"] or ids["iin"]:
                    match_key = find_matching_client(ids, clients, threshold)
                if match_key:
                    clients[match_key]["pages"].append(page)
                else:
                    remaining.append(page)
            if remaining:
                clients[primary_key]["pages"].extend(remaining)

    log.info(f"\n  Группировка завершена:")
    for key, data in clients.items():
        if key == "_unmatched":
            log.warning(f"    ⚠ Нераспознанные: {len(data['pages'])} стр.")
        else:
            phone_info = f" | тел: {data['phone']}" if data.get("phone") else ""
            iin_info = f" | ИИН: {data['iin']}" if data.get("iin") else ""
            pages_types = [p.get("page_type","?") for p in data["pages"]]
            log.info(f"    {data['name']}: {len(data['pages'])} стр. [{', '.join(pages_types)}]{phone_info}{iin_info}")

    return clients


# ============================================================
# 7. ДЕДУПЛИКАЦИЯ СТРАНИЦ
#    (убираем одинаковые карточки одного клиента)
# ============================================================

def compute_image_hash(image_path: str, hash_size: int = 16) -> str:
    """
    Перцептивный хэш изображения (average hash).
    Два одинаковых фото (даже разного размера/качества) дадут похожий хэш.
    """
    try:
        Image = _lazy_import_pil_image()
        img = Image.open(image_path).convert('L')  # Чёрно-белое
        img = img.resize((hash_size, hash_size), Image.Resampling.LANCZOS)
        pixels = list(img.getdata())
        avg = sum(pixels) / len(pixels)
        return ''.join('1' if p > avg else '0' for p in pixels)
    except Exception:
        return ""


def hamming_distance(hash1: str, hash2: str) -> float:
    """
    Расстояние Хэмминга между двумя хэшами (0.0 = идентичны, 1.0 = полностью разные).
    """
    if not hash1 or not hash2 or len(hash1) != len(hash2):
        return 1.0
    diff = sum(c1 != c2 for c1, c2 in zip(hash1, hash2))
    return diff / len(hash1)


def text_similarity(text1: str, text2: str) -> float:
    """Степень совпадения двух OCR-текстов (0.0 - 1.0)."""
    if not text1 or not text2:
        return 0.0
    t1 = " ".join(text1.lower().split())
    t2 = " ".join(text2.lower().split())
    if t1 == t2:
        return 1.0
    rf_fuzz = _lazy_import_rapidfuzz()
    if rf_fuzz is not None:
        return rf_fuzz.ratio(t1, t2) / 100.0
    return SequenceMatcher(None, t1, t2).ratio()


def deduplicate_pages(grouped_clients: dict) -> dict:
    """
    Убирает дубли страниц внутри каждого клиента.

    Три уровня обнаружения дублей:
    1. Перцептивный хэш фото — одно и то же фото (разный ракурс/качество)
    2. OCR-текст совпадает на 90%+ — одна и та же страница
    3. Тип + ключевые данные совпадают — одинаковое содержание

    При обнаружении дубля оставляет тот, у которого OCR-текст длиннее
    (т.е. более полно распознан).
    """
    hash_size = getattr(config, 'IMAGE_HASH_SIZE', 16)
    ocr_threshold = getattr(config, 'OCR_DUPLICATE_THRESHOLD', 0.90)
    total_removed = 0

    for key, client_data in grouped_clients.items():
        # Не трогаем непривязанные страницы, чтобы не потерять данные
        if key == "_unmatched":
            continue
        pages = client_data["pages"]
        if len(pages) <= 1:
            continue

        # Вычисляем хэши для всех страниц
        page_hashes = []
        for page in pages:
            filepath = page.get("filepath", "")
            img_hash = compute_image_hash(filepath, hash_size) if filepath and os.path.exists(filepath) else ""
            page_hashes.append(img_hash)

        # Ищем дубли
        to_remove = set()
        for i in range(len(pages)):
            if i in to_remove:
                continue
            for j in range(i + 1, len(pages)):
                if j in to_remove:
                    continue

                is_duplicate = False

                # Проверка 1: Перцептивный хэш фото (расстояние < 10% = дубль)
                if page_hashes[i] and page_hashes[j]:
                    distance = hamming_distance(page_hashes[i], page_hashes[j])
                    if distance < 0.10:
                        is_duplicate = True

                # Проверка 2: OCR-текст
                if not is_duplicate:
                    ocr_i = pages[i].get("ocr_text", "")
                    ocr_j = pages[j].get("ocr_text", "")
                    if ocr_i and ocr_j:
                        sim = text_similarity(ocr_i, ocr_j)
                        if sim >= ocr_threshold:
                            is_duplicate = True

                # Проверка 3: Тип + ключевые данные
                if not is_duplicate:
                    type_i = pages[i].get("page_type", "")
                    type_j = pages[j].get("page_type", "")
                    if type_i == type_j and type_i != "unknown":
                        data_i = json.dumps(pages[i].get("data", {}), sort_keys=True, ensure_ascii=False)
                        data_j = json.dumps(pages[j].get("data", {}), sort_keys=True, ensure_ascii=False)
                        if data_i == data_j:
                            is_duplicate = True

                if is_duplicate:
                    # Оставляем тот, у которого OCR-текст длиннее (лучше распознан)
                    len_i = len(pages[i].get("ocr_text", ""))
                    len_j = len(pages[j].get("ocr_text", ""))
                    remove_idx = j if len_i >= len_j else i
                    to_remove.add(remove_idx)

        if to_remove:
            removed_files = [pages[idx].get("filename", "?") for idx in to_remove]
            client_data["pages"] = [p for idx, p in enumerate(pages) if idx not in to_remove]
            total_removed += len(to_remove)
            name = client_data.get("name", key)
            log.info(f"    {name}: убрано {len(to_remove)} дубл. ({', '.join(removed_files)})")

    if total_removed:
        log.info(f"\n  Итого удалено дублей: {total_removed}")
    else:
        log.info(f"\n  Дублей не обнаружено")

    return grouped_clients


# ============================================================
# 8. ЗАПИСЬ В EXCEL
# ============================================================

def style_header(ws, row, col_count):
    styles = _get_excel_styles()
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = styles['HEADER_FONT']
        cell.fill = styles['HEADER_FILL']
        cell.alignment = styles['HEADER_ALIGNMENT']
        cell.border = styles['THIN_BORDER']


def style_data_cell(cell, warning=False):
    styles = _get_excel_styles()
    cell.font = styles['CELL_FONT']
    cell.alignment = styles['CELL_ALIGNMENT']
    cell.border = styles['THIN_BORDER']
    if warning:
        cell.fill = styles['WARN_FILL']


def auto_width(ws, min_w=12, max_w=50):
    from openpyxl.utils import get_column_letter
    for col_cells in ws.columns:
        max_len = 0
        letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, min_w), max_w)


def safe_val(d: dict, key: str, default=""):
    val = d.get(key, default)
    return val if val is not None else default


def truncate_text(text: str, max_length: int = 32000) -> str:
    """Ограничивает длину текста для Excel (лимит 32767 символов)."""
    if not text:
        return ""
    if len(text) <= max_length:
        return text
    return text[:max_length] + "... [ОБРЕЗАНО]"


def collect_ocr_texts(pages: list) -> dict:
    """
    Собирает OCR-тексты по типам страниц.

    Returns:
        dict с ключами: front, inner, procedures, products, complex, botox, full
    """
    texts = {
        "front": [],
        "inner": [],
        "procedures": [],
        "products": [],
        "complex": [],
        "botox": [],
    }

    for page in pages:
        page_type = page.get("page_type", "")
        ocr_text = page.get("ocr_text", "")

        if not ocr_text:
            continue

        # Маппинг типов страниц к категориям
        if page_type == "medical_card_front":
            texts["front"].append(ocr_text)
        elif page_type == "medical_card_inner":
            texts["inner"].append(ocr_text)
        elif page_type == "procedure_sheet":
            texts["procedures"].append(ocr_text)
        elif page_type == "products_list":
            texts["products"].append(ocr_text)
        elif page_type == "complex_package":
            texts["complex"].append(ocr_text)
        elif page_type == "botox_record":
            texts["botox"].append(ocr_text)
        else:
            # unknown → не теряем, кладём в полный текст и в procedures как fallback
            texts["procedures"].append(ocr_text)

    # Объединяем тексты по категориям
    result = {}
    for key, text_list in texts.items():
        joined = "\n\n---\n\n".join(text_list) if text_list else ""
        result[key] = truncate_text(joined)

    # Полный текст = все тексты вместе
    all_texts = []
    for text_list in texts.values():
        all_texts.extend(text_list)
    result["full"] = truncate_text("\n\n---\n\n".join(all_texts) if all_texts else "")

    # Если все целевые категории пусты, но full есть — дублируем в procedures как минимум,
    # чтобы колонка процедур не оставалась пустой.
    if all(not texts[key] for key in ["front", "inner", "procedures", "products", "complex", "botox"]) and result["full"]:
        result["procedures"] = result["full"]

    # Дополнительный дубль: если конкретная колонка пуста, но есть full — заполняем её full,
    # чтобы в Excel не оставались пустые OCR-колонки.
    if result["full"]:
        for key in ["front", "inner", "procedures", "products", "complex", "botox"]:
            if not result.get(key):
                result[key] = result["full"]

    # Собираем реконструированные таблицы
    tables_md_parts = []
    tables_csv_parts = []
    for page in pages:
        md = page.get("tables_md", "")
        csv_text = page.get("tables_csv", "")
        if md:
            tables_md_parts.append(md)
        if csv_text:
            tables_csv_parts.append(csv_text)

    result["tables_md"] = truncate_text("\n\n".join(tables_md_parts)) if tables_md_parts else ""
    result["tables_csv"] = truncate_text("\n\n".join(tables_csv_parts)) if tables_csv_parts else ""

    return result


def _build_client_row(key, cd, cid):
    """Собирает строку данных для листа Клиенты."""
    front = {}
    doctor = ""
    last_visit = ""
    first_visit = ""
    files = []

    for page in cd["pages"]:
        files.append(page.get("filename", ""))
        pt = page.get("page_type", "")
        d = page.get("data", {})

        if pt == "medical_card_front":
            front = d
        if d.get("doctor"):
            doctor = d["doctor"]
        if d.get("consultant"):
            doctor = doctor or d["consultant"]

        for proc_key in ["procedures", "products", "injections"]:
            items = d.get(proc_key, [])
            if isinstance(items, list):
                for item in items:
                    if isinstance(item, dict):
                        for date_key in ["date", "procedure_date"]:
                            dt = item.get(date_key, "")
                            if dt:
                                if not last_visit or dt > last_visit:
                                    last_visit = dt
                                if not first_visit or dt < first_visit:
                                    first_visit = dt

    photo_file = ""
    for page in cd["pages"]:
        if page.get("page_type") == "medical_card_front":
            photo_file = page.get("filename", "")
            break

    ocr_texts = collect_ocr_texts(cd["pages"])

    return [
        cid,
        safe_val(front, "card_created_date") or first_visit,
        photo_file,
        cd["name"],
        safe_val(front, "birth_date"),
        safe_val(front, "age"),
        safe_val(front, "gender"),
        safe_val(front, "citizenship"),
        cd.get("iin") or safe_val(front, "iin"),
        safe_val(front, "address"),
        cd.get("phone") or safe_val(front, "phone"),
        safe_val(front, "email"),
        safe_val(front, "messenger"),
        safe_val(front, "emergency_contact"),
        safe_val(front, "discount"),
        safe_val(front, "info_source"),
        safe_val(front, "allergies"),
        doctor, last_visit,
        len(cd["pages"]),
        "; ".join(files),
        ocr_texts.get("front", ""),
        ocr_texts.get("inner", ""),
        ocr_texts.get("procedures", ""),
        ocr_texts.get("products", ""),
        ocr_texts.get("complex", ""),
        ocr_texts.get("botox", ""),
        ocr_texts.get("full", ""),
        ocr_texts.get("tables_md", ""),
        ocr_texts.get("tables_csv", ""),
    ]


def _append_new_clients(grouped_clients: dict, output_path: str) -> int:
    """
    Дозаписывает ТОЛЬКО новых клиентов в существующий Excel.
    Существующие строки (с ручными правками) остаются нетронутыми.
    Новые клиенты определяются по файлам-источникам.

    Returns: количество добавленных клиентов (0 = нечего добавлять).
    Raises: Exception если дозапись невозможна (файл повреждён и т.п.)
    """
    import re
    from openpyxl import load_workbook

    wb = load_workbook(output_path)

    if "Клиенты" not in wb.sheetnames:
        wb.close()
        raise ValueError("Лист 'Клиенты' не найден")

    ws = wb["Клиенты"]

    # --- Читаем заголовки ---
    headers = [cell.value for cell in ws[1]]
    if "ID" not in headers or "Файлы-источники" not in headers:
        wb.close()
        raise ValueError("Нет колонок ID / Файлы-источники")

    id_col = headers.index("ID")
    files_col = headers.index("Файлы-источники")

    # --- Собираем файлы из существующих строк и макс. ID ---
    existing_files = set()
    max_id_num = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        cid_val = row[id_col] if id_col < len(row) else None
        files_val = row[files_col] if files_col < len(row) else None

        if cid_val:
            m = re.match(r'CL-(\d+)', str(cid_val))
            if m:
                max_id_num = max(max_id_num, int(m.group(1)))

        if files_val:
            for f in str(files_val).split("; "):
                f = f.strip()
                if f:
                    existing_files.add(f)

    # --- Определяем новых клиентов ---
    new_clients = {}
    for key, cd in grouped_clients.items():
        client_files = {p.get("filename", "") for p in cd["pages"]}
        client_files.discard("")
        if not client_files:
            continue
        # Клиент уже есть, если хотя бы один его файл в существующем Excel
        if client_files & existing_files:
            continue
        new_clients[key] = cd

    if not new_clients:
        wb.close()
        log.info(f"\n  Нет новых клиентов — файл сохранён без изменений.")
        return 0

    # --- Назначаем ID новым клиентам ---
    client_id_map = {}
    new_id_counter = max_id_num

    for key in sorted(new_clients.keys()):
        if key == "_unmatched":
            client_id_map[key] = "???"
        else:
            new_id_counter += 1
            client_id_map[key] = f"CL-{new_id_counter:04d}"

    # --- Дописываем в лист «Клиенты» ---
    row_idx = ws.max_row + 1

    for key in sorted(new_clients.keys()):
        cd = new_clients[key]
        cid = client_id_map[key]
        is_unmatched = key == "_unmatched"
        row_data = _build_client_row(key, cd, cid)

        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
            style_data_cell(ws.cell(row=row_idx, column=col_idx), warning=is_unmatched)

        row_idx += 1

    ws.auto_filter.ref = ws.dimensions

    # --- Дозапись в детальные листы ---
    # Структура: (имя_листа, заголовки, page_type, функция_извлечения_строк)

    def _ensure_sheet(wb, name, sheet_headers):
        if name in wb.sheetnames:
            return wb[name]
        ws_new = wb.create_sheet(name)
        ws_new.append(sheet_headers)
        style_header(ws_new, 1, len(sheet_headers))
        return ws_new

    # Мед_данные
    h_med = [
        "ID", "ФИО", "Основные жалобы", "Объективный статус",
        "Предварит. диагноз", "АД", "Вес", "ДМІ", "ДМІІ",
        "Охват груди", "Охват талии", "Охват бёдер",
        "Гепатиты/КВЗ/туберк./онко", "Хронические заболевания",
        "Отметки специалиста"
    ]
    ws_med = _ensure_sheet(wb, "Мед_данные", h_med)
    r_med = ws_med.max_row + 1

    for key in sorted(new_clients.keys()):
        cd = new_clients[key]
        cid = client_id_map[key]
        for page in cd["pages"]:
            if page.get("page_type") == "medical_card_inner":
                d = page.get("data", {})
                row_data = [
                    cid, cd["name"],
                    safe_val(d, "complaints"), safe_val(d, "objective_status"),
                    safe_val(d, "preliminary_diagnosis"),
                    safe_val(d, "blood_pressure"), safe_val(d, "weight"),
                    safe_val(d, "dm1"), safe_val(d, "dm2"),
                    safe_val(d, "chest"), safe_val(d, "waist"), safe_val(d, "hips"),
                    safe_val(d, "hepatitis_history"),
                    safe_val(d, "chronic_diseases"),
                    safe_val(d, "specialist_notes")
                ]
                for col_idx, val in enumerate(row_data, 1):
                    ws_med.cell(row=r_med, column=col_idx, value=val)
                    style_data_cell(ws_med.cell(row=r_med, column=col_idx))
                r_med += 1

    # Процедуры
    h_proc = ["ID", "ФИО", "Дата", "Процедура", "Описание", "Стоимость"]
    ws_proc = _ensure_sheet(wb, "Процедуры", h_proc)
    r_proc = ws_proc.max_row + 1

    for key in sorted(new_clients.keys()):
        cd = new_clients[key]
        cid = client_id_map[key]
        for page in cd["pages"]:
            if page.get("page_type") == "procedure_sheet":
                procs = page.get("data", {}).get("procedures", [])
                if isinstance(procs, list):
                    for p in procs:
                        if isinstance(p, dict):
                            row_data = [
                                cid, cd["name"],
                                safe_val(p, "date"),
                                safe_val(p, "procedure_name"),
                                safe_val(p, "description"),
                                safe_val(p, "cost")
                            ]
                            for col_idx, val in enumerate(row_data, 1):
                                ws_proc.cell(row=r_proc, column=col_idx, value=val)
                                style_data_cell(ws_proc.cell(row=r_proc, column=col_idx))
                            r_proc += 1

    # Покупки
    h_purch = ["ID", "ФИО", "Дата", "Консультант", "Наименование", "Цена"]
    ws_purch = _ensure_sheet(wb, "Покупки", h_purch)
    r_purch = ws_purch.max_row + 1

    for key in sorted(new_clients.keys()):
        cd = new_clients[key]
        cid = client_id_map[key]
        for page in cd["pages"]:
            if page.get("page_type") == "products_list":
                prods = page.get("data", {}).get("products", [])
                if isinstance(prods, list):
                    for p in prods:
                        if isinstance(p, dict):
                            row_data = [
                                cid, cd["name"],
                                safe_val(p, "date"),
                                safe_val(p, "consultant"),
                                safe_val(p, "product_name"),
                                safe_val(p, "price")
                            ]
                            for col_idx, val in enumerate(row_data, 1):
                                ws_purch.cell(row=r_purch, column=col_idx, value=val)
                                style_data_cell(ws_purch.cell(row=r_purch, column=col_idx))
                            r_purch += 1

    # Комплексы
    h_comp = [
        "ID", "Пациент", "Контакты", "Врач", "Комплекс",
        "Дата покупки", "Стоимость", "№", "Процедура",
        "Дата", "Кол-во", "Комментарий"
    ]
    ws_comp = _ensure_sheet(wb, "Комплексы", h_comp)
    r_comp = ws_comp.max_row + 1

    for key in sorted(new_clients.keys()):
        cd = new_clients[key]
        cid = client_id_map[key]
        for page in cd["pages"]:
            if page.get("page_type") == "complex_package":
                d = page.get("data", {})
                procs = d.get("procedures", [])
                base = [
                    cid, safe_val(d, "patient_name"), safe_val(d, "contacts"),
                    safe_val(d, "doctor"), safe_val(d, "complex_name"),
                    safe_val(d, "purchase_date"), safe_val(d, "complex_cost")
                ]
                if isinstance(procs, list) and procs:
                    for p in procs:
                        if isinstance(p, dict):
                            row_data = base + [
                                safe_val(p, "number"), safe_val(p, "procedure"),
                                safe_val(p, "date"), safe_val(p, "quantity"),
                                safe_val(p, "comment")
                            ]
                            for col_idx, val in enumerate(row_data, 1):
                                ws_comp.cell(row=r_comp, column=col_idx, value=val)
                                style_data_cell(ws_comp.cell(row=r_comp, column=col_idx))
                            r_comp += 1
                else:
                    row_data = base + ["", "", "", "", ""]
                    for col_idx, val in enumerate(row_data, 1):
                        ws_comp.cell(row=r_comp, column=col_idx, value=val)
                        style_data_cell(ws_comp.cell(row=r_comp, column=col_idx))
                    r_comp += 1

    # Ботокс
    h_bot = [
        "ID", "ФИО", "Препарат", "Область введения",
        "Кол-во единиц", "Общая доза", "Дата процедуры", "Дата контроля"
    ]
    ws_bot = _ensure_sheet(wb, "Ботокс", h_bot)
    r_bot = ws_bot.max_row + 1

    for key in sorted(new_clients.keys()):
        cd = new_clients[key]
        cid = client_id_map[key]
        for page in cd["pages"]:
            if page.get("page_type") == "botox_record":
                d = page.get("data", {})
                injs = d.get("injections", [])
                if isinstance(injs, list):
                    for inj in injs:
                        if isinstance(inj, dict):
                            row_data = [
                                cid, cd["name"],
                                safe_val(inj, "drug"),
                                safe_val(inj, "injection_area"),
                                safe_val(inj, "units_count"),
                                safe_val(inj, "total_dose"),
                                safe_val(inj, "procedure_date"),
                                safe_val(inj, "control_date")
                            ]
                            for col_idx, val in enumerate(row_data, 1):
                                ws_bot.cell(row=r_bot, column=col_idx, value=val)
                                style_data_cell(ws_bot.cell(row=r_bot, column=col_idx))
                            r_bot += 1

    # --- Сохранение ---
    wb.save(output_path)
    wb.close()

    log.info(f"\n  ✓ Дозаписано {len(new_clients)} новых клиентов в {output_path}")
    return len(new_clients)


def write_to_excel(grouped_clients: dict, all_results: list):
    from openpyxl import Workbook

    output_path = config.OUTPUT_FILE
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)

    # === РЕЖИМ ДОЗАПИСИ (если файл уже существует) ===
    if os.path.exists(output_path):
        try:
            _append_new_clients(grouped_clients, output_path)
            return
        except Exception as e:
            log.warning(f"  ⚠ Дозапись не удалась ({e}), пересоздаю файл...")

    # === СОЗДАНИЕ С НУЛЯ ===
    wb = Workbook()

    # === ЛИСТ 1: КЛИЕНТЫ ===
    ws = wb.active
    ws.title = "Клиенты"
    headers = [
        # 3.1. Идентификационный блок
        "ID", "Дата создания карты", "Фото (файл)",
        # 3.2. Персональные данные
        "ФИО", "Дата рождения", "Возраст", "Пол", "Гражданство",
        "ИИН / Паспорт", "Адрес", "Телефон", "Email", "Мессенджер",
        "Экстренный контакт",
        # Дополнительные поля клиники
        "Скидка", "Источник инфо", "Аллергии", "Консультант/Врач",
        "Дата последнего визита", "Кол-во страниц", "Файлы-источники",
        # 3.3. OCR-тексты (оцифрованный текст)
        "OCR_Текст_Лицевая", "OCR_Текст_Внутренняя", "OCR_Текст_Процедуры",
        "OCR_Текст_Покупки", "OCR_Текст_Комплексы", "OCR_Текст_Ботокс",
        "OCR_Текст_Полный",
        # 3.4. Реконструированные таблицы
        "OCR_Таблицы_MD", "OCR_Таблицы_CSV"
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    client_id_map = {}
    row_idx = 2

    for idx, (key, cd) in enumerate(
        sorted(grouped_clients.items(), key=lambda x: x[0]), 1
    ):
        is_unmatched = key == "_unmatched"
        cid = "???" if is_unmatched else f"CL-{idx:04d}"
        client_id_map[key] = cid

        front = {}
        doctor = ""
        last_visit = ""
        first_visit = ""
        files = []

        for page in cd["pages"]:
            files.append(page.get("filename", ""))
            pt = page.get("page_type", "")
            d = page.get("data", {})

            if pt == "medical_card_front":
                front = d
            if d.get("doctor"):
                doctor = d["doctor"]
            if d.get("consultant"):
                doctor = doctor or d["consultant"]

            # Находим последнюю и первую дату визита
            for proc_key in ["procedures", "products", "injections"]:
                items = d.get(proc_key, [])
                if isinstance(items, list):
                    for item in items:
                        if isinstance(item, dict):
                            for date_key in ["date", "procedure_date"]:
                                dt = item.get(date_key, "")
                                if dt:
                                    if not last_visit or dt > last_visit:
                                        last_visit = dt
                                    if not first_visit or dt < first_visit:
                                        first_visit = dt

        # Фото — первый файл с лицевой стороной
        photo_file = ""
        for page in cd["pages"]:
            if page.get("page_type") == "medical_card_front":
                photo_file = page.get("filename", "")
                break

        # Собираем OCR-тексты
        ocr_texts = collect_ocr_texts(cd["pages"])

        row = [
            # 3.1. Идентификационный блок
            cid,
            safe_val(front, "card_created_date") or first_visit,
            photo_file,
            # 3.2. Персональные данные (всё из OCR, пустое = пустое)
            cd["name"],
            safe_val(front, "birth_date"),
            safe_val(front, "age"),
            safe_val(front, "gender"),
            safe_val(front, "citizenship"),
            cd.get("iin") or safe_val(front, "iin"),
            safe_val(front, "address"),
            cd.get("phone") or safe_val(front, "phone"),
            safe_val(front, "email"),
            safe_val(front, "messenger"),
            safe_val(front, "emergency_contact"),
            # Дополнительные поля клиники
            safe_val(front, "discount"),
            safe_val(front, "info_source"),
            safe_val(front, "allergies"),
            doctor, last_visit,
            len(cd["pages"]),
            "; ".join(files),
            # OCR-тексты
            ocr_texts.get("front", ""),
            ocr_texts.get("inner", ""),
            ocr_texts.get("procedures", ""),
            ocr_texts.get("products", ""),
            ocr_texts.get("complex", ""),
            ocr_texts.get("botox", ""),
            ocr_texts.get("full", ""),
            # Реконструированные таблицы
            ocr_texts.get("tables_md", ""),
            ocr_texts.get("tables_csv", ""),
        ]
        ws.append(row)
        for col in range(1, len(headers) + 1):
            style_data_cell(ws.cell(row=row_idx, column=col), warning=is_unmatched)
        row_idx += 1

    auto_width(ws)
    ws.auto_filter.ref = ws.dimensions

    # === ЛИСТ 2: МЕД. ДАННЫЕ ===
    ws2 = wb.create_sheet("Мед_данные")
    h2 = [
        "ID", "ФИО", "Основные жалобы", "Объективный статус",
        "Предварит. диагноз", "АД", "Вес", "ДМІ", "ДМІІ",
        "Охват груди", "Охват талии", "Охват бёдер",
        "Гепатиты/КВЗ/туберк./онко", "Хронические заболевания",
        "Отметки специалиста"
    ]
    ws2.append(h2)
    style_header(ws2, 1, len(h2))
    r2 = 2

    for key, cd in sorted(grouped_clients.items()):
        cid = client_id_map.get(key, "")
        for page in cd["pages"]:
            if page.get("page_type") == "medical_card_inner":
                d = page.get("data", {})
                ws2.append([
                    cid, cd["name"],
                    safe_val(d, "complaints"), safe_val(d, "objective_status"),
                    safe_val(d, "preliminary_diagnosis"),
                    safe_val(d, "blood_pressure"), safe_val(d, "weight"),
                    safe_val(d, "dm1"), safe_val(d, "dm2"),
                    safe_val(d, "chest"), safe_val(d, "waist"), safe_val(d, "hips"),
                    safe_val(d, "hepatitis_history"),
                    safe_val(d, "chronic_diseases"),
                    safe_val(d, "specialist_notes")
                ])
                for col in range(1, len(h2) + 1):
                    style_data_cell(ws2.cell(row=r2, column=col))
                r2 += 1

    auto_width(ws2)
    ws2.auto_filter.ref = ws2.dimensions

    # === ЛИСТ 3: ПРОЦЕДУРЫ ===
    ws3 = wb.create_sheet("Процедуры")
    h3 = ["ID", "ФИО", "Дата", "Процедура", "Описание", "Стоимость"]
    ws3.append(h3)
    style_header(ws3, 1, len(h3))
    r3 = 2

    for key, cd in sorted(grouped_clients.items()):
        cid = client_id_map.get(key, "")
        for page in cd["pages"]:
            if page.get("page_type") == "procedure_sheet":
                procs = page.get("data", {}).get("procedures", [])
                if isinstance(procs, list):
                    for p in procs:
                        if isinstance(p, dict):
                            ws3.append([
                                cid, cd["name"],
                                safe_val(p, "date"),
                                safe_val(p, "procedure_name"),
                                safe_val(p, "description"),
                                safe_val(p, "cost")
                            ])
                            for col in range(1, len(h3) + 1):
                                style_data_cell(ws3.cell(row=r3, column=col))
                            r3 += 1

    auto_width(ws3)
    ws3.auto_filter.ref = ws3.dimensions

    # === ЛИСТ 4: ПОКУПКИ ===
    ws4 = wb.create_sheet("Покупки")
    h4 = ["ID", "ФИО", "Дата", "Консультант", "Наименование", "Цена"]
    ws4.append(h4)
    style_header(ws4, 1, len(h4))
    r4 = 2

    for key, cd in sorted(grouped_clients.items()):
        cid = client_id_map.get(key, "")
        for page in cd["pages"]:
            if page.get("page_type") == "products_list":
                prods = page.get("data", {}).get("products", [])
                if isinstance(prods, list):
                    for p in prods:
                        if isinstance(p, dict):
                            ws4.append([
                                cid, cd["name"],
                                safe_val(p, "date"),
                                safe_val(p, "consultant"),
                                safe_val(p, "product_name"),
                                safe_val(p, "price")
                            ])
                            for col in range(1, len(h4) + 1):
                                style_data_cell(ws4.cell(row=r4, column=col))
                            r4 += 1

    auto_width(ws4)
    ws4.auto_filter.ref = ws4.dimensions

    # === ЛИСТ 5: КОМПЛЕКСЫ ===
    ws5 = wb.create_sheet("Комплексы")
    h5 = [
        "ID", "Пациент", "Контакты", "Врач", "Комплекс",
        "Дата покупки", "Стоимость", "№", "Процедура",
        "Дата", "Кол-во", "Комментарий"
    ]
    ws5.append(h5)
    style_header(ws5, 1, len(h5))
    r5 = 2

    for key, cd in sorted(grouped_clients.items()):
        cid = client_id_map.get(key, "")
        for page in cd["pages"]:
            if page.get("page_type") == "complex_package":
                d = page.get("data", {})
                procs = d.get("procedures", [])
                base = [
                    cid, safe_val(d, "patient_name"), safe_val(d, "contacts"),
                    safe_val(d, "doctor"), safe_val(d, "complex_name"),
                    safe_val(d, "purchase_date"), safe_val(d, "complex_cost")
                ]
                if isinstance(procs, list) and procs:
                    for p in procs:
                        if isinstance(p, dict):
                            ws5.append(base + [
                                safe_val(p, "number"), safe_val(p, "procedure"),
                                safe_val(p, "date"), safe_val(p, "quantity"),
                                safe_val(p, "comment")
                            ])
                            for col in range(1, len(h5) + 1):
                                style_data_cell(ws5.cell(row=r5, column=col))
                            r5 += 1
                else:
                    ws5.append(base + ["", "", "", "", ""])
                    for col in range(1, len(h5) + 1):
                        style_data_cell(ws5.cell(row=r5, column=col))
                    r5 += 1

    auto_width(ws5)
    ws5.auto_filter.ref = ws5.dimensions

    # === ЛИСТ 6: БОТОКС ===
    ws6 = wb.create_sheet("Ботокс")
    h6 = [
        "ID", "ФИО", "Препарат", "Область введения",
        "Кол-во единиц", "Общая доза", "Дата процедуры", "Дата контроля"
    ]
    ws6.append(h6)
    style_header(ws6, 1, len(h6))
    r6 = 2

    for key, cd in sorted(grouped_clients.items()):
        cid = client_id_map.get(key, "")
        for page in cd["pages"]:
            if page.get("page_type") == "botox_record":
                d = page.get("data", {})
                injs = d.get("injections", [])
                if isinstance(injs, list):
                    for inj in injs:
                        if isinstance(inj, dict):
                            ws6.append([
                                cid, cd["name"],
                                safe_val(inj, "drug"),
                                safe_val(inj, "injection_area"),
                                safe_val(inj, "units_count"),
                                safe_val(inj, "total_dose"),
                                safe_val(inj, "procedure_date"),
                                safe_val(inj, "control_date")
                            ])
                            for col in range(1, len(h6) + 1):
                                style_data_cell(ws6.cell(row=r6, column=col))
                            r6 += 1

    auto_width(ws6)
    ws6.auto_filter.ref = ws6.dimensions

    # === СОХРАНЕНИЕ ===
    os.makedirs(os.path.dirname(config.OUTPUT_FILE) or '.', exist_ok=True)
    wb.save(config.OUTPUT_FILE)

    log.info(f"\nExcel сохранён: {config.OUTPUT_FILE}")
    log.info(f"\nСтатистика:")
    log.info(f"  Клиентов: {ws.max_row - 1}")
    log.info(f"  Мед. записей: {ws2.max_row - 1}")
    log.info(f"  Процедур: {ws3.max_row - 1}")
    log.info(f"  Покупок: {ws4.max_row - 1}")
    log.info(f"  Комплексов: {ws5.max_row - 1}")
    log.info(f"  Записей ботокса: {ws6.max_row - 1}")


# ============================================================
# 9. MAIN
# ============================================================

def main():
    log.info("=" * 60)
    log.info("  ОЦИФРОВКА КАРТОЧЕК КЛИЕНТОВ")
    log.info("  Google Vision + Claude API → Excel")
    log.info("=" * 60)

    # Конфигурация в лог
    log.debug(f"[КОНФИГ] Vision: {config.GOOGLE_VISION_CREDENTIALS}")
    log.debug(f"[КОНФИГ] Модель: {config.CLAUDE_MODEL}")
    log.debug(f"[КОНФИГ] Папка фото: {config.INPUT_FOLDER}")
    log.debug(f"[КОНФИГ] Выход: {config.OUTPUT_FILE}")
    log.debug(f"[КОНФИГ] Fuzzy порог: {getattr(config, 'FUZZY_NAME_THRESHOLD', 0.75)}")
    log.debug(f"[КОНФИГ] Dedup порог: {getattr(config, 'OCR_DUPLICATE_THRESHOLD', 0.90)}")

    # Проверки
    if not os.path.exists(config.GOOGLE_VISION_CREDENTIALS):
        log.error(f"Файл Google Vision не найден: {config.GOOGLE_VISION_CREDENTIALS}")
        sys.exit(1)

    if "ваш-ключ" in config.ANTHROPIC_API_KEY:
        log.error("Укажите Claude API ключ (vibeproxy) в config.py → ANTHROPIC_API_KEY")
        sys.exit(1)

    if not os.path.exists(config.INPUT_FOLDER):
        log.error(f"Папка не найдена: {config.INPUT_FOLDER}")
        sys.exit(1)

    base_url = getattr(config, 'ANTHROPIC_BASE_URL', None)
    if base_url:
        log.info(f"  Proxy: {base_url}")

    log.info("\nИнициализация API...")
    vision_client = init_vision_client()
    claude_client = init_claude_client()
    log.info("  Google Vision — OK")
    log.info("  Claude API — OK")

    # Обработка
    t_start = time.time()
    results = process_all_images(vision_client, claude_client)

    # Сырые результаты
    raw_path = os.path.join(os.path.dirname(config.OUTPUT_FILE) or '.', "raw_results.json")
    with open(raw_path, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    log.info(f"\nСырые данные: {raw_path}")

    # Группировка
    log.info("\nГруппировка по клиентам (fuzzy matching)...")
    grouped = group_by_client(results)
    log.info(f"  Уникальных клиентов: {len([k for k in grouped if k != '_unmatched'])}")

    # Дедупликация
    log.info("\nПоиск дублей страниц...")
    grouped = deduplicate_pages(grouped)

    # Excel
    log.info("\nЗапись в Excel...")
    write_to_excel(grouped, results)

    elapsed = time.time() - t_start
    log.info(f"\n{'═' * 60}")
    log.info(f"  ГОТОВО! Время: {elapsed:.0f}с ({elapsed/60:.1f} мин)")
    log.info(f"{'═' * 60}")


if __name__ == "__main__":
    main()
